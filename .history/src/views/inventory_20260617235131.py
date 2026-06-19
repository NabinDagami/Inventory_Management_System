import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import sys
import os
import io
import random
import string
import tempfile
import threading
from PIL import Image
from barcode import Code128
from barcode.writer import ImageWriter

# Add parent directory to path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from models.database import db
from utils.sku_generator import SKUGenerator
import utils.simple_table_styles as table_styles
from utils.export_manager import ExportManager
# from utils.camera_scanner import CameraBarcodeReader
from utils.barcode_server import BarcodeServer


class _FilterPopup(ctk.CTkToplevel):
    """CTkToplevel that skips the Windows titlebar-color withdraw/deiconify cycle.
    This avoids grab_set() being called on a withdrawn window and prevents
    the bad-window-path-name crashes from after() callbacks firing on a destroyed popup."""
    _deactivate_windows_window_header_manipulation = True

    def destroy(self):
        try:
            self.grab_release()
        except Exception:
            pass
        super().destroy()


class ScrollableFilter(ctk.CTkFrame):
    """Dropdown with search, checkmarks, and scroll support."""

    def __init__(self, master, values=None, variable=None, command=None,
                 width=140, **kwargs):
        super().__init__(master, fg_color="transparent", **kwargs)

        self._values = values or ["All"]
        self._variable = variable
        self._command = command
        self._width = width
        self._popup = None
        self._filtered = list(self._values)

        initial = self._variable.get() if self._variable else "All"
        self._is_active = initial != "All"

        self._btn = ctk.CTkButton(
            self, text=initial, width=width, height=28,
            anchor="w",
            fg_color=("gray90", "gray25"),
            text_color=("gray10", "gray90"),
            hover_color=("gray85", "gray35"),
            corner_radius=6,
            font=ctk.CTkFont(size=12),
            command=self._open_popup
        )
        self._btn.pack(fill="both", expand=True)

        if self._variable:
            self._variable.trace_add("write", lambda *_: self._on_var_changed())

    def _on_var_changed(self):
        val = self._variable.get()
        self._btn.configure(text=val)
        was = self._is_active
        self._is_active = val != "All"
        if was != self._is_active:
            self._update_active_style()

    def _update_active_style(self):
        if self._is_active:
            self._btn.configure(
                fg_color=("#DBEAFE", "#1E3A5F"),
                text_color=("#1E40AF", "#93C5FD"),
                border_color=("#3B82F6", "#60A5FA"),
                border_width=1
            )
        else:
            self._btn.configure(
                fg_color=("gray90", "gray25"),
                text_color=("gray10", "gray90"),
                border_width=0
            )

    def _open_popup(self):
        self._close_popup()

        current = self._variable.get() if self._variable else "All"
        btn_font = self._btn.cget("font")
        popup_w = self._width
        for v in self._values:
            tw = btn_font.measure(f"✓ {v}") + 50
            if tw > popup_w:
                popup_w = min(tw, 400)

        n_items = len(self._values)
        max_h = min(300, n_items * 32 + 50)

        # Close other filter popups before opening a new one
        for child in self.master.winfo_children():
            if isinstance(child, ScrollableFilter) and child is not self and child._popup:
                child._close_popup()

        popup = _FilterPopup(self)
        popup.withdraw()
        popup.overrideredirect(True)
        popup.attributes("-topmost", True)
        # NO transient() and NO grab_set() — they block all input to the parent
        # window on Windows. Instead _on_global_click closes the popup when
        # clicking outside, and the click falls through to the intended widget.

        main = ctk.CTkFrame(popup, corner_radius=8,
                            fg_color=("white", "#262626"),
                            border_width=1, border_color=("gray70", "gray50"))
        main.pack(fill="both", expand=True)

        # Search bar
        search_frame = ctk.CTkFrame(main, fg_color="transparent")
        search_frame.pack(fill="x", padx=6, pady=(6, 2))
        search_entry = ctk.CTkEntry(
            search_frame, height=28, corner_radius=6,
            placeholder_text="Type to filter...",
            font=ctk.CTkFont(size=12),
        )
        search_entry.pack(fill="x")

        sep = ctk.CTkFrame(main, height=1, fg_color=("gray80", "gray40"))
        sep.pack(fill="x", padx=6, pady=4)

        # Scrollable options
        scroll = ctk.CTkScrollableFrame(
            main, corner_radius=0,
            fg_color="transparent",
            scrollbar_button_hover_color=("gray70", "gray45")
        )
        scroll.pack(fill="both", expand=True, padx=2, pady=(0, 4))
        scroll.configure(width=popup_w)

        self._filtered = list(self._values)
        self._option_btns = []

        def rebuild_list(*_):
            q = search_entry.get().lower()
            self._filtered = [v for v in self._values if q in v.lower()]
            for ob in self._option_btns:
                ob.pack_forget()
            self._option_btns.clear()
            for value in self._filtered:
                selected = (value == current)
                prefix = "✓ " if selected else "   "
                ob = ctk.CTkButton(
                    scroll, text=f"{prefix}{value}", height=30, anchor="w",
                    fg_color=("#EFF6FF", "#1E3A5F") if selected else "transparent",
                    text_color=("#1E40AF", "#93C5FD") if selected else ("gray10", "gray90"),
                    hover_color=("#F3F4F6", "#333333"),
                    corner_radius=4,
                    font=ctk.CTkFont(size=13),
                    command=lambda v=value: self._select(v)
                )
                ob.pack(fill="x", padx=4, pady=1)
                self._option_btns.append(ob)

        search_entry.bind("<KeyRelease>", rebuild_list)
        rebuild_list()

        # Show popup before setting geometry
        popup.deiconify()
        popup.update_idletasks()

        sx = self.winfo_rootx()
        sy = self.winfo_rooty() + self._btn.winfo_height()
        popup.geometry(f"{popup_w}x{max_h}+{int(sx)}+{int(sy)}")

        self._popup = popup
        popup.bind("<Escape>", lambda e: self._close_popup())
        popup.after(10, lambda: self._safe_focus(popup, search_entry))
        toplevel = self.winfo_toplevel()
        toplevel.bind("<Button-1>", self._on_global_click, add="+")

    def _safe_focus(self, popup, entry):
        try:
            if popup.winfo_exists():
                popup.focus_force()
                entry.focus_set()
        except Exception:
            pass

    def _close_popup(self):
        if getattr(self, '_closing_popup', False):
            return
        self._closing_popup = True
        try:
            # ALWAYS release the grab first — even if the popup is gone,
            # releasing defensively prevents stuck-grab lock on the toplevel
            try:
                toplevel = self.winfo_toplevel()
                toplevel.grab_release()
            except Exception:
                pass
            if not self._popup:
                self._closing_popup = False
                return
            popup = self._popup
            self._popup = None
            try:
                if popup.winfo_exists():
                    popup.destroy()
            except Exception:
                pass
        finally:
            self._closing_popup = False

    def _select(self, value):
        if self._variable:
            self._variable.set(value)
        self._btn.configure(text=value)
        self._is_active = value != "All"
        self._update_active_style()
        if self._command:
            self._command(value)
        self._close_popup()

    def _on_global_click(self, event):
        """Called when any click happens on the main toplevel.
        Close the popup if the click was not inside it or the filter button."""
        if not self._popup:
            return
        # Check if click is inside the popup
        try:
            px = self._popup.winfo_rootx()
            py = self._popup.winfo_rooty()
            pw = self._popup.winfo_width()
            ph = self._popup.winfo_height()
            in_popup = (px <= event.x_root <= px + pw and py <= event.y_root <= py + ph)
        except Exception:
            in_popup = False

        # Check if click is on the filter button
        try:
            bx = self._btn.winfo_rootx()
            by = self._btn.winfo_rooty()
            bw = self._btn.winfo_width()
            bh = self._btn.winfo_height()
            in_btn = (bx <= event.x_root <= bx + bw and by <= event.y_root <= by + bh)
        except Exception:
            in_btn = False

        if not in_popup and not in_btn:
            self._close_popup()

    def configure(self, **kwargs):
        if "values" in kwargs:
            self._values = kwargs.pop("values")
        if "command" in kwargs:
            self._command = kwargs.pop("command")
        if kwargs:
            super().configure(**kwargs)


class InventoryView:
    def __init__(self, parent):
        self.parent = parent
        self.current_tab = "products"
        self.create_inventory_view()
        self.load_data()
    
    def create_inventory_view(self):
        """Create inventory management interface"""
        # Main container
        main_frame = ctk.CTkFrame(self.parent)
        main_frame.pack(fill="both", expand=True, padx=15, pady=15)
        
        # Header with tabs and action buttons
        header_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        header_frame.pack(fill="x", padx=5, pady=(8, 0))
        
        # Tab buttons — pill-style segmented control
        tab_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
        tab_frame.pack(side="left", padx=10, pady=10)
        
        tab_bg = ctk.CTkFrame(tab_frame, corner_radius=20, fg_color=("gray90", "gray20"))
        tab_bg.pack(side="left")
        
        tab_defs = [
            ("products", "Products", 100),
            ("categories", "Categories", 110),
            ("sub_categories", "Sub Categories", 130),
            ("brands", "Brands", 100),
            ("sub_brands", "Sub Brands", 120),
        ]
        self._tab_buttons = {}
        for i, (key, text, w) in enumerate(tab_defs):
            pad = (3, 0) if i > 0 else (3, 3)
            btn = ctk.CTkButton(
                tab_bg,
                text=text,
                width=w,
                height=32,
                corner_radius=16,
                font=ctk.CTkFont(size=12, weight="bold"),
                fg_color=("gray85", "gray25"),
                hover_color=("gray75", "gray35"),
                text_color=("gray10", "gray90"),
                command=lambda k=key: self.switch_tab(k)
            )
            btn.pack(side="left", padx=pad, pady=3)
            self._tab_buttons[key] = btn
        self.products_btn = self._tab_buttons["products"]
        self.categories_btn = self._tab_buttons["categories"]
        self.sub_categories_btn = self._tab_buttons["sub_categories"]
        self.brands_btn = self._tab_buttons["brands"]
        self.sub_brands_btn = self._tab_buttons["sub_brands"]
        
        # Action buttons - Visual Hierarchy
        action_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
        action_frame.pack(side="right", padx=10, pady=10)
        
        # Primary Action: Add Product (Blue)
        self.add_btn = ctk.CTkButton(
            action_frame,
            text="  + Add Product",
            width=130,
            height=36,
            font=ctk.CTkFont(size=12, weight="bold"),
            command=self.add_item,
            fg_color="#3B82F6",
            hover_color="#2563EB",
            corner_radius=8
        )
        self.add_btn.pack(side="left", padx=(0, 6))
        
        # Secondary: Edit (Outline)
        self.edit_btn = ctk.CTkButton(
            action_frame,
            text="✏  Edit",
            width=80,
            height=36,
            font=ctk.CTkFont(size=12, weight="bold"),
            command=self.edit_item,
            fg_color="transparent",
            border_color="#6B7280",
            border_width=2,
            text_color=("gray10", "gray90"),
            hover_color=("gray85", "gray25"),
            corner_radius=8
        )
        self.edit_btn.pack(side="left", padx=(0, 6))
        
        # Danger: Delete (Red)
        self.delete_btn = ctk.CTkButton(
            action_frame,
            text="🗑  Delete",
            width=90,
            height=36,
            font=ctk.CTkFont(size=12, weight="bold"),
            command=self.delete_item,
            fg_color="#EF4444",
            hover_color="#DC2626",
            corner_radius=8
        )
        self.delete_btn.pack(side="left", padx=(6, 12))
        
        # Separator
        sep = ctk.CTkFrame(action_frame, width=1, height=28, fg_color=("gray75", "gray40"))
        sep.pack(side="left", padx=(0, 12))
        
        # Export Dropdown
        self.export_menu = ctk.CTkOptionMenu(
            action_frame,
            values=["📥 Export", "📊  Excel", "📄  PDF"],
            width=110,
            height=36,
            font=ctk.CTkFont(size=11, weight="bold"),
            command=self.handle_export,
            fg_color=("gray80", "gray30"),
            button_color=("gray70", "gray40"),
            button_hover_color=("gray60", "gray50"),
            text_color=("gray20", "gray90"),
            corner_radius=8
        )
        self.export_menu.pack(side="left")
        self.export_menu.set("📥 Export")
        
        # Import button
        self.import_btn = ctk.CTkButton(
            action_frame,
            text="📥  Import",
            width=100,
            height=36,
            font=ctk.CTkFont(size=12, weight="bold"),
            command=self.import_from_excel,
            fg_color="#10B981",
            hover_color="#059669",
            corner_radius=8
        )
        self.import_btn.pack(side="left", padx=(8, 0))
        
        # Search and Filter frame
        self.filter_frame = ctk.CTkFrame(main_frame, corner_radius=10)
        self.filter_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        # Row 1: Search box
        search_row = ctk.CTkFrame(self.filter_frame, fg_color="transparent")
        search_row.pack(fill="x", padx=12, pady=(10, 5))
        
        search_icon = ctk.CTkLabel(search_row, text="🔍", font=ctk.CTkFont(size=14))
        search_icon.pack(side="left", padx=(0, 6))
        
        self.search_var = ctk.StringVar()
        self.search_var.trace('w', self.filter_data)
        search_entry = ctk.CTkEntry(
            search_row,
            textvariable=self.search_var,
            placeholder_text="Search by name, SKU or barcode...",
            width=320,
            height=32,
            corner_radius=8
        )
        search_entry.pack(side="left", padx=5)
        
        # Separator line
        sep = ctk.CTkFrame(self.filter_frame, height=1, fg_color=("gray80", "gray35"))
        sep.pack(fill="x", padx=12, pady=(5, 5))
        
        # Row 2: Advanced filters - inline layout
        filters_row = ctk.CTkFrame(self.filter_frame, fg_color="transparent")
        filters_row.pack(fill="x", padx=10, pady=(5, 10))
        
        # Filter label prefix
        filter_label = ctk.CTkLabel(filters_row, text="Filters:", font=ctk.CTkFont(size=11, weight="bold"))
        filter_label.pack(side="left", padx=(2, 8))
        
        # Category filter
        cat_label = ctk.CTkLabel(filters_row, text="Category", font=ctk.CTkFont(size=10), text_color=("gray50", "gray60"))
        cat_label.pack(side="left", padx=(0, 3))
        
        self.category_filter_var = tk.StringVar(value="All")
        self.category_filter = ScrollableFilter(
            filters_row,
            variable=self.category_filter_var,
            values=["All"],
            width=130,
            command=lambda v: self.filter_data()
        )
        self.category_filter.pack(side="left", padx=(0, 10))
        
        # Sub Category filter
        sub_cat_label = ctk.CTkLabel(filters_row, text="Sub Cat", font=ctk.CTkFont(size=10), text_color=("gray50", "gray60"))
        sub_cat_label.pack(side="left", padx=(0, 3))
        
        self.sub_category_filter_var = tk.StringVar(value="All")
        self.sub_category_filter = ScrollableFilter(
            filters_row,
            variable=self.sub_category_filter_var,
            values=["All"],
            width=130,
            command=lambda v: self.filter_data()
        )
        self.sub_category_filter.pack(side="left", padx=(0, 10))
        
        # Brand filter
        brand_label = ctk.CTkLabel(filters_row, text="Brand", font=ctk.CTkFont(size=10), text_color=("gray50", "gray60"))
        brand_label.pack(side="left", padx=(0, 3))
        
        self.brand_filter_var = tk.StringVar(value="All")
        self.brand_filter = ScrollableFilter(
            filters_row,
            variable=self.brand_filter_var,
            values=["All"],
            width=130,
            command=lambda v: self.filter_data()
        )
        self.brand_filter.pack(side="left", padx=(0, 10))
        
        # Sub Brand filter
        sub_brand_label = ctk.CTkLabel(filters_row, text="Sub Brand", font=ctk.CTkFont(size=10), text_color=("gray50", "gray60"))
        sub_brand_label.pack(side="left", padx=(0, 3))
        
        self.sub_brand_filter_var = tk.StringVar(value="All")
        self.sub_brand_filter = ScrollableFilter(
            filters_row,
            variable=self.sub_brand_filter_var,
            values=["All"],
            width=130,
            command=lambda v: self.filter_data()
        )
        self.sub_brand_filter.pack(side="left", padx=(0, 10))
        
        # Stock status filter
        stock_label = ctk.CTkLabel(filters_row, text="Stock", font=ctk.CTkFont(size=10), text_color=("gray50", "gray60"))
        stock_label.pack(side="left", padx=(0, 3))
        
        self.stock_filter_var = tk.StringVar(value="All")
        self.stock_filter = ScrollableFilter(
            filters_row,
            variable=self.stock_filter_var,
            values=["All", "In Stock", "Low Stock", "Out of Stock"],
            width=110,
            command=lambda v: self.filter_data()
        )
        self.stock_filter.pack(side="left", padx=(0, 10))
        
        # Clear filters button - inline with filters
        clear_btn = ctk.CTkButton(
            filters_row,
            text="✕  Clear",
            command=self.clear_filters,
            width=85,
            height=28,
            fg_color=("gray80", "gray30"),
            hover_color=("gray70", "gray40"),
            text_color=("gray20", "gray90"),
            font=ctk.CTkFont(size=11),
            corner_radius=8
        )
        clear_btn.pack(side="left")
        
        # Active filter chips row (hidden initially, shown when filters active)
        self.filter_chips_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        self.filter_chips_inner = ctk.CTkFrame(self.filter_chips_frame, fg_color="transparent")
        self.filter_chips_inner.pack(fill="x", padx=12, pady=(0, 4))
        
        # Summary cards row
        self.create_summary_cards(main_frame)
        
        # Content area with table
        content_frame = ctk.CTkFrame(main_frame)
        content_frame.pack(fill="both", expand=True, padx=5, pady=(0, 5))
        
        # Create table
        self.create_table(content_frame)
    
    def create_summary_cards(self, parent):
        """Create compact summary cards showing inventory statistics"""
        self.cards_frame = ctk.CTkFrame(parent, fg_color="transparent")
        self.cards_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        card_configs = [
            ("📦", "Total Products", ("#3B82F6", "#2563EB"), self.show_all_products),
            ("⚡", "Low Stock", ("#D97706", "#B45309"), lambda: self.filter_by_stock_status("Low Stock")),
            ("❌", "Out of Stock", ("#DC2626", "#B91C1C"), lambda: self.filter_by_stock_status("Out of Stock")),
            ("💰", "Inventory Value", ("#059669", "#047857"), None),
        ]
        
        labels = {}
        for i, (icon, text, color, cmd) in enumerate(card_configs):
            card = ctk.CTkFrame(
                self.cards_frame, corner_radius=10,
                fg_color=color, cursor="hand2" if cmd else "arrow",
                height=64, border_width=0
            )
            card.pack(side="left", fill="x", expand=True, padx=(0 if i == 3 else 0, 8 if i < 3 else 0))
            card.pack_propagate(False)
            if cmd:
                card.bind("<Button-1>", lambda e, c=cmd: c())
            inner = ctk.CTkFrame(card, fg_color="transparent")
            inner.pack(expand=True, fill="both", padx=14, pady=8)
            inner.grid_columnconfigure(0, weight=0)
            inner.grid_columnconfigure(1, weight=1)
            ctk.CTkLabel(
                inner, text=icon,
                font=ctk.CTkFont(size=20), text_color="white"
            ).grid(row=0, column=0, rowspan=2, padx=(0, 10))
            ctk.CTkLabel(
                inner, text=text,
                font=ctk.CTkFont(size=10), text_color=("#FFFFFF", "#D0D0D0")
            ).grid(row=0, column=1, sticky="w")
            val_lbl = ctk.CTkLabel(
                inner, text="0",
                font=ctk.CTkFont(size=16, weight="bold"), text_color="white"
            )
            val_lbl.grid(row=1, column=1, sticky="w")
            key = text.lower().replace(" ", "_")
            labels[key] = val_lbl
        
        self.total_label = labels["total_products"]
        self.low_stock_label = labels["low_stock"]
        self.out_stock_label = labels["out_of_stock"]
        self.value_label = labels["inventory_value"]
    
    def show_all_products(self):
        """Show all products by clearing filters"""
        self.clear_filters()
    
    def filter_by_stock_status(self, status):
        """Filter table by stock status"""
        self.stock_filter_var.set(status)
        self.filter_data()
    
    def _cleanup_scrollbars(self):
        """Clean up scrollbar references before destroying treeview to prevent TclError"""
        try:
            # Reset scrollbar commands to prevent calls to destroyed widgets
            if hasattr(self, 'h_scrollbar') and self.h_scrollbar and self.h_scrollbar.winfo_exists():
                self.h_scrollbar.configure(command=None)
            if hasattr(self, 'v_scrollbar') and self.v_scrollbar and self.v_scrollbar.winfo_exists():
                self.v_scrollbar.configure(command=None)
            # Reset treeview scroll commands
            if hasattr(self, 'data_tree') and self.data_tree and self.data_tree.winfo_exists():
                self.data_tree.configure(yscrollcommand=None, xscrollcommand=None)
        except Exception:
            # Ignore any errors during cleanup
            pass
    
    def create_table(self, parent):
        """Create the data table"""
        # Table frame
        table_frame = ctk.CTkFrame(parent)
        table_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Products table (default)
        self.create_products_table(table_frame)
    def refresh_table_tags(self):
        """Re-apply zebra striping, stock status tags, and context menu colors for current theme."""
        if not hasattr(self, 'data_tree') or not self.data_tree:
            return
        if not self.data_tree.winfo_exists():
            return
        is_dark = ctk.get_appearance_mode() == "Dark"
        self.data_tree.tag_configure("evenrow", background="#F8FAFC" if not is_dark else "#1A1A2E")
        self.data_tree.tag_configure("oddrow",  background="#FFFFFF" if not is_dark else "#1E1E2E")
        self.data_tree.tag_configure("stock_zero", foreground=("#DC2626" if not is_dark else "#F87171"), font=("Segoe UI", 11, "bold"))
        self.data_tree.tag_configure("stock_low",  foreground=("#D97706" if not is_dark else "#FBBF24"), font=("Segoe UI", 11, "bold"))
        self.data_tree.tag_configure("stock_ok",   foreground=("#059669" if not is_dark else "#34D399"), font=("Segoe UI", 11))
        # Update context menu colors
        if hasattr(self, 'context_menu') and self.context_menu:
            try:
                ctx_bg = "#2D2D2D" if is_dark else "#F8FAFC"
                ctx_fg = "#FFFFFF" if is_dark else "#1F2937"
                self.context_menu.configure(bg=ctx_bg, fg=ctx_fg)
            except Exception:
                pass

    
    def create_products_table(self, parent):
        """Create products table with enhanced styling and smooth animations"""
        # Clear existing table - first unconfigure scrollbars to prevent errors
        self._cleanup_scrollbars()
        for widget in parent.winfo_children():
            widget.destroy()
        
        # Create treeview frame with clean border
        table_container = ctk.CTkFrame(parent, border_width=1, border_color=("gray80", "gray40"))
        table_container.pack(fill="both", expand=True, padx=0, pady=0)
        table_container.grid_columnconfigure(0, weight=1)
        table_container.grid_rowconfigure(0, weight=1)
        
        # Create treeview for products
        columns = ("#", "SKU", "Name", "Category", "Sub Category", "Brand", "Sub Brand", "Stock", "Qty Sold", "Normal Price", "Workshop Price", "Reorder Level")
        self.data_tree = ttk.Treeview(table_container, columns=columns, show="headings", height=20)
        self.data_tree.grid(row=0, column=0, sticky="nsew", padx=(2, 0), pady=(2, 0))
        
        # Apply centralized styling
        table_styles.apply_product_style(self.data_tree)
        
        # Alternating row colors
        is_dark = ctk.get_appearance_mode() == "Dark"
        self.data_tree.tag_configure("evenrow", background="#F8FAFC" if not is_dark else "#1A1A2E")
        self.data_tree.tag_configure("oddrow",  background="#FFFFFF" if not is_dark else "#1E1E2E")
        
        # Stock status tag colors (adapt to dark mode)
        self.data_tree.tag_configure("stock_zero", foreground=("#DC2626" if not is_dark else "#F87171"), font=("Segoe UI", 11, "bold"))
        self.data_tree.tag_configure("stock_low", foreground=("#D97706" if not is_dark else "#FBBF24"), font=("Segoe UI", 11, "bold"))
        self.data_tree.tag_configure("stock_ok", foreground=("#059669" if not is_dark else "#34D399"), font=("Segoe UI", 11))
        
        # Column widths - Name and Brand stretch to fill space
        column_widths = {"#": 50, "SKU": 110, "Name": 200, "Category": 120, "Sub Category": 120,
                        "Brand": 120, "Sub Brand": 120, "Stock": 80, "Qty Sold": 80,
                        "Normal Price": 100, "Workshop Price": 120, "Reorder Level": 90}
        
        for col in columns:
            self.data_tree.heading(col, text=f"  {col}  ", anchor="center")
            stretch = col in ("Name", "Category", "Brand")
            anchor = "center" if col != "#" else "w"
            self.data_tree.column(col, width=column_widths.get(col, 120), anchor=anchor, minwidth=70, stretch=stretch)
        
        # Scrollbars (both inside table_container, using grid)
        self.v_scrollbar = ttk.Scrollbar(table_container, orient="vertical", command=self.data_tree.yview)
        self.v_scrollbar.grid(row=0, column=1, sticky="ns", pady=(2, 0))
        
        self.h_scrollbar = ttk.Scrollbar(table_container, orient="horizontal", command=self.data_tree.xview)
        self.h_scrollbar.grid(row=1, column=0, columnspan=2, sticky="ew", padx=(2, 0))
        
        self.data_tree.configure(yscrollcommand=self.v_scrollbar.set, xscrollcommand=self.h_scrollbar.set)
        
        # Bind double-click to edit
        self.data_tree.bind("<Double-1>", lambda e: self.edit_item())
        
        # Bind right-click for context menu
        self.data_tree.bind("<Button-3>", self.show_context_menu)
        
        # Create styled context menu (adapts to theme)
        is_dark = ctk.get_appearance_mode() == "Dark"
        ctx_bg = "#2D2D2D" if is_dark else "#F8FAFC"
        ctx_fg = "#FFFFFF" if is_dark else "#1F2937"
        self.context_menu = tk.Menu(self.parent, tearoff=0, bg=ctx_bg, fg=ctx_fg, 
                                    activebackground="#3B82F6", activeforeground="#FFFFFF",
                                    borderwidth=1, relief="solid",
                                    font=("Segoe UI", 11))
        self.context_menu.add_command(label="  ✏  Edit", command=self.edit_item)
        self.context_menu.add_command(label="  ➕  Quick Stock +1", command=lambda: self.quick_stock_adjust(1))
        self.context_menu.add_command(label="  ➖  Quick Stock -1", command=lambda: self.quick_stock_adjust(-1))
        self.context_menu.add_separator()
        self.context_menu.add_command(label="  🗑  Delete", command=self.delete_item, foreground="#EF4444")
        
        # Item count label below table
        self.count_label = ctk.CTkLabel(parent, text="Showing 0 of 0 items", font=ctk.CTkFont(size=11), text_color=("gray50", "gray60"))
        self.count_label.pack(side="bottom", anchor="w", padx=5, pady=(5, 5))
    
    def show_context_menu(self, event):
        """Show right-click context menu"""
        # Select row under mouse
        row_id = self.data_tree.identify_row(event.y)
        if row_id:
            self.data_tree.selection_set(row_id)
            self.context_menu.post(event.x_root, event.y_root)
    
    def quick_stock_adjust(self, amount):
        """Quickly adjust stock by +/- 1"""
        selection = self.data_tree.selection()
        if not selection:
            return
        
        item_values = self.data_tree.item(selection[0], 'values')
        sku = item_values[1]
        
        # Get current stock
        product = db.execute_query("SELECT stock FROM products WHERE sku = ?", (sku,))
        if not product:
            return
        
        current_stock = product[0]['stock']
        new_stock = current_stock + amount
        
        # Validate - prevent negative stock
        if new_stock < 0:
            messagebox.showerror("Invalid Operation", "Stock cannot be negative!")
            return
        
        # Update stock
        try:
            db.execute_update("UPDATE products SET stock = ? WHERE sku = ?", (new_stock, sku))
            self.load_data()
            messagebox.showinfo("Success", f"Stock updated! New stock: {new_stock}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update stock: {e}")
    
    def create_categories_table(self, parent):
        """Create categories table with enhanced styling"""
        self._cleanup_scrollbars()
        for widget in parent.winfo_children():
            widget.destroy()
        
        columns = ("ID", "Name", "Description", "Created")
        self.data_tree = ttk.Treeview(parent, columns=columns, show="headings", height=25)
        table_styles.apply_category_style(self.data_tree)
        
        is_dark = ctk.get_appearance_mode() == "Dark"
        self.data_tree.tag_configure("evenrow", background="#F8FAFC" if not is_dark else "#1A1A2E")
        self.data_tree.tag_configure("oddrow",  background="#FFFFFF" if not is_dark else "#1E1E2E")
        
        column_widths = {"ID": 80, "Name": 250, "Description": 450, "Created": 180}
        for col in columns:
            self.data_tree.heading(col, text=f"  {col}  ", anchor="center")
            self.data_tree.column(col, width=column_widths.get(col, 120), anchor="center" if col == "ID" else "w", minwidth=80)
        
        self.v_scrollbar = ttk.Scrollbar(parent, orient="vertical", command=self.data_tree.yview)
        self.h_scrollbar = ttk.Scrollbar(parent, orient="horizontal", command=self.data_tree.xview)
        self.data_tree.configure(yscrollcommand=self.v_scrollbar.set, xscrollcommand=self.h_scrollbar.set)
        
        self.data_tree.pack(side="left", fill="both", expand=True, padx=(5, 0), pady=5)
        self.v_scrollbar.pack(side="right", fill="y", pady=5)
        self.h_scrollbar.pack(side="bottom", fill="x", padx=5)
        self.data_tree.bind("<Double-1>", lambda e: self.edit_item())
    
    def create_sub_categories_table(self, parent):
        """Create sub categories table with enhanced styling"""
        self._cleanup_scrollbars()
        for widget in parent.winfo_children():
            widget.destroy()
        
        columns = ("ID", "Name", "Category", "Description", "Created")
        self.data_tree = ttk.Treeview(parent, columns=columns, show="headings", height=25)
        table_styles.apply_category_style(self.data_tree)
        
        is_dark = ctk.get_appearance_mode() == "Dark"
        self.data_tree.tag_configure("evenrow", background="#F8FAFC" if not is_dark else "#1A1A2E")
        self.data_tree.tag_configure("oddrow",  background="#FFFFFF" if not is_dark else "#1E1E2E")
        
        column_widths = {"ID": 80, "Name": 200, "Category": 150, "Description": 350, "Created": 180}
        for col in columns:
            self.data_tree.heading(col, text=f"  {col}  ", anchor="center")
            self.data_tree.column(col, width=column_widths.get(col, 120), anchor="center" if col == "ID" else "w", minwidth=80)
        
        self.v_scrollbar = ttk.Scrollbar(parent, orient="vertical", command=self.data_tree.yview)
        self.h_scrollbar = ttk.Scrollbar(parent, orient="horizontal", command=self.data_tree.xview)
        self.data_tree.configure(yscrollcommand=self.v_scrollbar.set, xscrollcommand=self.h_scrollbar.set)
        
        self.data_tree.pack(side="left", fill="both", expand=True, padx=(5, 0), pady=5)
        self.v_scrollbar.pack(side="right", fill="y", pady=5)
        self.h_scrollbar.pack(side="bottom", fill="x", padx=5)
        self.data_tree.bind("<Double-1>", lambda e: self.edit_item())
    
    def create_brands_table(self, parent):
        """Create brands table with enhanced styling"""
        self._cleanup_scrollbars()
        for widget in parent.winfo_children():
            widget.destroy()
        
        columns = ("ID", "Name", "Description", "Created")
        self.data_tree = ttk.Treeview(parent, columns=columns, show="headings", height=25)
        table_styles.apply_brand_style(self.data_tree)
        
        is_dark = ctk.get_appearance_mode() == "Dark"
        self.data_tree.tag_configure("evenrow", background="#F8FAFC" if not is_dark else "#1A1A2E")
        self.data_tree.tag_configure("oddrow",  background="#FFFFFF" if not is_dark else "#1E1E2E")
        
        column_widths = {"ID": 80, "Name": 250, "Description": 450, "Created": 180}
        for col in columns:
            self.data_tree.heading(col, text=f"  {col}  ", anchor="center")
            self.data_tree.column(col, width=column_widths.get(col, 120), anchor="center" if col == "ID" else "w", minwidth=80)
        
        self.v_scrollbar = ttk.Scrollbar(parent, orient="vertical", command=self.data_tree.yview)
        self.h_scrollbar = ttk.Scrollbar(parent, orient="horizontal", command=self.data_tree.xview)
        self.data_tree.configure(yscrollcommand=self.v_scrollbar.set, xscrollcommand=self.h_scrollbar.set)
        
        self.data_tree.pack(side="left", fill="both", expand=True, padx=(5, 0), pady=5)
        self.v_scrollbar.pack(side="right", fill="y", pady=5)
        self.h_scrollbar.pack(side="bottom", fill="x", padx=5)
        self.data_tree.bind("<Double-1>", lambda e: self.edit_item())
    
    def create_sub_brands_table(self, parent):
        """Create sub brands table with enhanced styling"""
        self._cleanup_scrollbars()
        for widget in parent.winfo_children():
            widget.destroy()
        
        columns = ("ID", "Name", "Brand", "Description", "Created")
        self.data_tree = ttk.Treeview(parent, columns=columns, show="headings", height=25)
        table_styles.apply_brand_style(self.data_tree)
        
        is_dark = ctk.get_appearance_mode() == "Dark"
        self.data_tree.tag_configure("evenrow", background="#F8FAFC" if not is_dark else "#1A1A2E")
        self.data_tree.tag_configure("oddrow",  background="#FFFFFF" if not is_dark else "#1E1E2E")
        
        column_widths = {"ID": 80, "Name": 200, "Brand": 150, "Description": 350, "Created": 180}
        for col in columns:
            self.data_tree.heading(col, text=f"  {col}  ", anchor="center")
            self.data_tree.column(col, width=column_widths.get(col, 120), anchor="center" if col == "ID" else "w", minwidth=80)
        
        self.v_scrollbar = ttk.Scrollbar(parent, orient="vertical", command=self.data_tree.yview)
        self.h_scrollbar = ttk.Scrollbar(parent, orient="horizontal", command=self.data_tree.xview)
        self.data_tree.configure(yscrollcommand=self.v_scrollbar.set, xscrollcommand=self.h_scrollbar.set)
        
        self.data_tree.pack(side="left", fill="both", expand=True, padx=(5, 0), pady=5)
        self.v_scrollbar.pack(side="right", fill="y", pady=5)
        self.h_scrollbar.pack(side="bottom", fill="x", padx=5)
        self.data_tree.bind("<Double-1>", lambda e: self.edit_item())
    
    def switch_tab(self, tab_name):
        """Switch between tabs"""
        self.current_tab = tab_name
        
        active_fg = ("#3B82F6", "#2563EB")
        active_text = "white"
        inactive_fg = ("gray85", "gray25")
        inactive_text = ("gray10", "gray90")
        
        for key, btn in self._tab_buttons.items():
            is_active = key == tab_name
            btn.configure(
                fg_color=active_fg if is_active else inactive_fg,
                text_color=active_text if is_active else inactive_text,
                hover_color=("#2563EB", "#1D4ED8") if is_active else ("gray75", "gray35")
            )
        
        # Update add button text
        button_texts = {
            "products": "Add Product",
            "categories": "Add Category",
            "sub_categories": "Add Sub Category",
            "brands": "Add Brand",
            "sub_brands": "Add Sub Brand"
        }
        self.add_btn.configure(text=button_texts.get(tab_name, "Add"))
        
        # Close any open filter popups and hide filter bar for non-products tabs
        if tab_name != "products":
            for f in (self.category_filter, self.sub_category_filter,
                      self.brand_filter, self.sub_brand_filter,
                      self.stock_filter):
                f._close_popup()
            self.filter_frame.pack_forget()
        else:
            self.filter_frame.pack(fill="x", padx=10, pady=(0, 10), before=self.cards_frame)
        
        # Recreate table for current tab
        content_frame = self.data_tree.master
        if tab_name == "products":
            self.create_products_table(content_frame)
        elif tab_name == "categories":
            self.create_categories_table(content_frame)
        elif tab_name == "sub_categories":
            self.create_sub_categories_table(content_frame)
        elif tab_name == "brands":
            self.create_brands_table(content_frame)
        elif tab_name == "sub_brands":
            self.create_sub_brands_table(content_frame)
        
        # Load data
        self.load_data()
    
    def load_data(self):
        """Load data for current tab"""
        try:
            # Clear existing items
            for item in self.data_tree.get_children():
                self.data_tree.delete(item)
            
            if self.current_tab == "products":
                self.load_products_data()
            elif self.current_tab == "categories":
                self.load_categories_data()
            elif self.current_tab == "sub_categories":
                self.load_sub_categories_data()
            elif self.current_tab == "brands":
                self.load_brands_data()
            elif self.current_tab == "sub_brands":
                self.load_sub_brands_data()
                
            if hasattr(self, 'filter_chips_frame'):
                self._update_filter_chips()
                
        except Exception as e:
            print(f"Error loading data: {e}")
            messagebox.showerror("Error", f"Failed to load data: {e}")
    
    def format_price(self, price):
        """Format price with comma separators"""
        from utils.format_utils import format_price as _fp
        return _fp(price)
    
    def get_stock_status_tag(self, stock, reorder_level):
        """Get stock status color tag - only for stock column"""
        if stock == 0:
            return "stock_zero"  # Red
        elif stock <= reorder_level:
            return "stock_low"   # Orange/Yellow
        else:
            return "stock_ok"    # Green
    
    def load_products_data(self):
        """Load products data with proper formatting, stock colors, and summary stats"""
        query = """
            SELECT p.sku, p.name, c.category_name, sc.sub_category_name, 
                   b.brand_name, sb.sub_brand_name, p.stock, p.qty_sold,
                   p.price_normal, p.price_workshop, p.reorder_level, p.cost_price,
                   p.barcode
            FROM products p
            LEFT JOIN categories c ON p.category_id = c.category_id
            LEFT JOIN sub_categories sc ON p.sub_category_id = sc.sub_category_id
            LEFT JOIN brands b ON p.brand_id = b.brand_id
            LEFT JOIN sub_brands sb ON p.sub_brand_id = sb.sub_brand_id
            WHERE p.is_active = 1
            ORDER BY p.name
        """
        products = db.execute_query(query)
        
        self.all_data = products  # Store for filtering
        
        # Calculate summary statistics
        total_products = len(products)
        low_stock_count = sum(1 for p in products if 0 < p['stock'] <= p['reorder_level'])
        out_stock_count = sum(1 for p in products if p['stock'] == 0)
        inventory_value = sum(p['stock'] * p['cost_price'] for p in products)
        
        # Update summary cards
        self.total_label.configure(text=str(total_products))
        self.low_stock_label.configure(text=str(low_stock_count))
        self.out_stock_label.configure(text=str(out_stock_count))
        self.value_label.configure(text=self.format_price(inventory_value))
        
        # Load filter options
        self.load_filter_options()
        
        # Insert data with proper formatting
        for idx, product in enumerate(products, start=1):
            # Stock display with status icon and label
            stock = product['stock']
            reorder = product['reorder_level']
            if stock == 0:
                stock_display = "0  Out"
            elif stock <= reorder:
                stock_display = f"{stock}  Low"
            else:
                stock_display = f"{stock}  OK"
            
            # Get stock color tag
            stock_tag = self.get_stock_status_tag(stock, reorder)
            
            # Insert the item with alternating row and stock status tags
            row_tag = "evenrow" if idx % 2 == 0 else "oddrow"
            item_id = self.data_tree.insert("", "end", values=(
                idx,
                product['sku'],
                product['name'],
                product['category_name'] or "N/A",
                product['sub_category_name'] or "-",
                product['brand_name'] or "N/A",
                product['sub_brand_name'] or "-",
                stock_display,
                product.get('qty_sold', 0),
                self.format_price(product['price_normal']),
                self.format_price(product['price_workshop']),
                reorder
            ), tags=(stock_tag, row_tag))
        
        displayed_count = len(self.data_tree.get_children())
        if hasattr(self, 'count_label'):
            self.count_label.configure(text=f"Showing {displayed_count} of {total_products} Products")
    
    def load_categories_data(self):
        """Load categories data"""
        categories = db.execute_query("SELECT * FROM categories ORDER BY category_name")
        self.all_data = categories
        
        for idx, category in enumerate(categories, start=1):
            row_tag = "evenrow" if idx % 2 == 0 else "oddrow"
            self.data_tree.insert("", "end", values=(
                category['category_id'],
                category['category_name'],
                category['description'] or "N/A",
                category['created_at']
            ), tags=(row_tag,))
    
    def load_sub_categories_data(self):
        """Load sub categories data"""
        query = """
            SELECT sc.*, c.category_name 
            FROM sub_categories sc
            LEFT JOIN categories c ON sc.category_id = c.category_id
            ORDER BY sc.sub_category_name
        """
        sub_categories = db.execute_query(query)
        self.all_data = sub_categories
        
        for idx, sc in enumerate(sub_categories, start=1):
            row_tag = "evenrow" if idx % 2 == 0 else "oddrow"
            self.data_tree.insert("", "end", values=(
                sc['sub_category_id'],
                sc['sub_category_name'],
                sc['category_name'] or "N/A",
                sc['description'] or "N/A",
                sc['created_at']
            ), tags=(row_tag,))
    
    def load_brands_data(self):
        """Load brands data"""
        brands = db.execute_query("SELECT * FROM brands ORDER BY brand_name")
        self.all_data = brands
        
        for idx, brand in enumerate(brands, start=1):
            row_tag = "evenrow" if idx % 2 == 0 else "oddrow"
            self.data_tree.insert("", "end", values=(
                brand['brand_id'],
                brand['brand_name'],
                brand['description'] or "N/A",
                brand['created_at']
            ), tags=(row_tag,))
    
    def load_sub_brands_data(self):
        """Load sub brands data"""
        query = """
            SELECT sb.*, b.brand_name 
            FROM sub_brands sb
            LEFT JOIN brands b ON sb.brand_id = b.brand_id
            ORDER BY sb.sub_brand_name
        """
        sub_brands = db.execute_query(query)
        self.all_data = sub_brands
        
        for idx, sb in enumerate(sub_brands, start=1):
            row_tag = "evenrow" if idx % 2 == 0 else "oddrow"
            self.data_tree.insert("", "end", values=(
                sb['sub_brand_id'],
                sb['sub_brand_name'],
                sb['brand_name'] or "N/A",
                sb['description'] or "N/A",
                sb['created_at']
            ), tags=(row_tag,))
    
    def load_filter_options(self):
        """Load categories and brands for filter dropdowns"""
        # Load categories
        categories = db.execute_query("SELECT category_name FROM categories ORDER BY category_name")
        cat_values = ["All"] + [cat['category_name'] for cat in categories if cat['category_name']]
        self.category_filter.configure(values=cat_values)
        
        # Load sub categories
        sub_categories = db.execute_query("SELECT sub_category_name FROM sub_categories ORDER BY sub_category_name")
        sub_cat_values = ["All"] + [sc['sub_category_name'] for sc in sub_categories if sc['sub_category_name']]
        self.sub_category_filter.configure(values=sub_cat_values)
        
        # Load brands
        brands = db.execute_query("SELECT brand_name FROM brands ORDER BY brand_name")
        brand_values = ["All"] + [brand['brand_name'] for brand in brands if brand['brand_name']]
        self.brand_filter.configure(values=brand_values)
        
        # Load sub brands
        sub_brands = db.execute_query("SELECT sub_brand_name FROM sub_brands ORDER BY sub_brand_name")
        sub_brand_values = ["All"] + [sb['sub_brand_name'] for sb in sub_brands if sb['sub_brand_name']]
        self.sub_brand_filter.configure(values=sub_brand_values)
    
    def clear_filters(self):
        """Clear all filters and reload data"""
        self.search_var.set("")
        self.category_filter_var.set("All")
        self.sub_category_filter_var.set("All")
        self.brand_filter_var.set("All")
        self.sub_brand_filter_var.set("All")
        self.stock_filter_var.set("All")
        self.filter_data()
    
    def _update_filter_chips(self):
        for w in self.filter_chips_inner.winfo_children():
            w.destroy()
        filters = {}
        if self.category_filter_var.get() != "All":
            filters["Category"] = self.category_filter_var.get()
        if self.sub_category_filter_var.get() != "All":
            filters["Sub Cat"] = self.sub_category_filter_var.get()
        if self.brand_filter_var.get() != "All":
            filters["Brand"] = self.brand_filter_var.get()
        if self.sub_brand_filter_var.get() != "All":
            filters["Sub Brand"] = self.sub_brand_filter_var.get()
        if self.stock_filter_var.get() != "All":
            filters["Stock"] = self.stock_filter_var.get()
        search = self.search_var.get().strip()
        if search:
            filters["Search"] = f'"{search}"'
        if not filters:
            self.filter_chips_frame.pack_forget()
            return
        lbl = ctk.CTkLabel(
            self.filter_chips_inner, text="Active:",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color=("gray50", "gray60")
        )
        lbl.pack(side="left", padx=(0, 6))
        for key, val in filters.items():
            chip = ctk.CTkFrame(
                self.filter_chips_inner, corner_radius=12,
                fg_color=("#DBEAFE", "#1E3A5F"),
                border_width=0
            )
            ctk.CTkLabel(
                chip, text=f"{key}: {val}",
                font=ctk.CTkFont(size=10),
                text_color=("#1E40AF", "#93C5FD")
            ).pack(side="left", padx=(8, 2), pady=2)
            rm_btn = ctk.CTkLabel(
                chip, text=" ✕",
                font=ctk.CTkFont(size=10, weight="bold"),
                text_color=("#1E40AF", "#93C5FD"),
                cursor="hand2"
            )
            rm_btn.pack(side="right", padx=(2, 6), pady=2)
            def on_rm(e, k=key):
                if k == "Search":
                    self.search_var.set("")
                elif k == "Category":
                    self.category_filter_var.set("All")
                elif k == "Sub Cat":
                    self.sub_category_filter_var.set("All")
                elif k == "Brand":
                    self.brand_filter_var.set("All")
                elif k == "Sub Brand":
                    self.sub_brand_filter_var.set("All")
                elif k == "Stock":
                    self.stock_filter_var.set("All")
            rm_btn.bind("<Button-1>", on_rm)
            chip.pack(side="left", padx=(0, 5))
        # Pack before the summary cards to maintain layout order
        self.filter_chips_frame.pack(fill="x", padx=10, pady=(0, 4), before=self.cards_frame)
    
    def filter_data(self, *args):
        """Filter data based on search and filter criteria"""
        search_term = self.search_var.get().lower()
        selected_category = self.category_filter_var.get()
        selected_sub_category = self.sub_category_filter_var.get()
        selected_brand = self.brand_filter_var.get()
        selected_sub_brand = self.sub_brand_filter_var.get()
        selected_stock = self.stock_filter_var.get()
        
        # Clear tree
        for item in self.data_tree.get_children():
            self.data_tree.delete(item)
        
        if not hasattr(self, 'all_data'):
            return
        
        # Filter and display matching items
        row_num = 0
        for item in self.all_data:
            if self.current_tab == "products":
                # Text search filter
                matches_search = (search_term in item['name'].lower() or 
                                  search_term in item['sku'].lower() or
                                  (item.get('barcode') and search_term in item['barcode'].lower()))
                
                # Category filter
                category_name = item['category_name'] if item['category_name'] else "N/A"
                matches_category = (selected_category == "All" or 
                                    category_name == selected_category)
                
                # Sub Category filter
                sub_category_name = item['sub_category_name'] if item['sub_category_name'] else "N/A"
                matches_sub_category = (selected_sub_category == "All" or 
                                        sub_category_name == selected_sub_category)
                
                # Brand filter
                brand_name = item['brand_name'] if item['brand_name'] else "N/A"
                matches_brand = (selected_brand == "All" or 
                                 brand_name == selected_brand)
                
                # Sub Brand filter
                sub_brand_name = item['sub_brand_name'] if item['sub_brand_name'] else "N/A"
                matches_sub_brand = (selected_sub_brand == "All" or 
                                     sub_brand_name == selected_sub_brand)
                
                # Stock status filter
                stock = item['stock']
                reorder_level = item['reorder_level']
                if selected_stock == "All":
                    matches_stock = True
                elif selected_stock == "In Stock":
                    matches_stock = stock > reorder_level
                elif selected_stock == "Low Stock":
                    matches_stock = 0 < stock <= reorder_level
                elif selected_stock == "Out of Stock":
                    matches_stock = stock == 0
                else:
                    matches_stock = True
                
                # Apply all filters
                if matches_search and matches_category and matches_sub_category and matches_brand and matches_sub_brand and matches_stock:
                    row_num += 1
                    # Stock display with status icon and label
                    if stock == 0:
                        stock_display = "0  Out"
                    elif stock <= reorder_level:
                        stock_display = f"{stock}  Low"
                    else:
                        stock_display = f"{stock}  OK"
                    
                    # Get stock color tag
                    stock_tag = self.get_stock_status_tag(stock, reorder_level)
                    
                    # Insert the item with alternating row and stock status tags
                    row_tag = "evenrow" if row_num % 2 == 0 else "oddrow"
                    item_id = self.data_tree.insert("", "end", values=(
                        row_num,
                        item['sku'],
                        item['name'],
                        category_name,
                        sub_category_name,
                        brand_name,
                        sub_brand_name,
                        stock_display,
                        item.get('qty_sold', 0),
                        self.format_price(item['price_normal']),
                        self.format_price(item['price_workshop']),
                        reorder_level
                    ), tags=(stock_tag, row_tag))
            elif self.current_tab == "categories":
                if search_term in item['category_name'].lower():
                    row_num += 1
                    row_tag = "evenrow" if row_num % 2 == 0 else "oddrow"
                    self.data_tree.insert("", "end", values=(
                        item['category_id'],
                        item['category_name'],
                        item['description'] or "N/A",
                        item['created_at']
                    ), tags=(row_tag,))
            elif self.current_tab == "sub_categories":
                if search_term in item['sub_category_name'].lower():
                    row_num += 1
                    row_tag = "evenrow" if row_num % 2 == 0 else "oddrow"
                    self.data_tree.insert("", "end", values=(
                        item['sub_category_id'],
                        item['sub_category_name'],
                        item['category_name'] or "N/A",
                        item['description'] or "N/A",
                        item['created_at']
                    ), tags=(row_tag,))
            elif self.current_tab == "brands":
                if search_term in item['brand_name'].lower():
                    row_num += 1
                    row_tag = "evenrow" if row_num % 2 == 0 else "oddrow"
                    self.data_tree.insert("", "end", values=(
                        item['brand_id'],
                        item['brand_name'],
                        item['description'] or "N/A",
                        item['created_at']
                    ), tags=(row_tag,))
            elif self.current_tab == "sub_brands":
                if search_term in item['sub_brand_name'].lower():
                    row_num += 1
                    row_tag = "evenrow" if row_num % 2 == 0 else "oddrow"
                    self.data_tree.insert("", "end", values=(
                        item['sub_brand_id'],
                        item['sub_brand_name'],
                        item['brand_name'] or "N/A",
                        item['description'] or "N/A",
                        item['created_at']
                    ), tags=(row_tag,))
        
        # Update count label
        displayed_count = len(self.data_tree.get_children())
        if hasattr(self, 'count_label'):
            total = len(self.all_data) if hasattr(self, 'all_data') else 0
            tab_name = self.current_tab.replace("_", " ").title()
            self.count_label.configure(text=f"Showing {displayed_count} of {total} {tab_name}")
        
        # Update filter chips
        if hasattr(self, 'filter_chips_frame'):
            self._update_filter_chips()
    
    def add_item(self):
        """Add new item based on current tab"""
        if self.current_tab == "products":
            self.add_product()
        elif self.current_tab == "categories":
            self.add_category()
        elif self.current_tab == "sub_categories":
            self.add_sub_category()
        elif self.current_tab == "brands":
            self.add_brand()
        elif self.current_tab == "sub_brands":
            self.add_sub_brand()
    
    def add_product(self):
        """Add new product"""
        dialog = ProductDialog(self.parent, "Add Product")
        if dialog.result:
            try:
                # Generate SKU
                sku = SKUGenerator.generate_sku(dialog.result['category_id'], dialog.result['brand_id'])
                
                # Insert product
                db.execute_insert(
                    """INSERT INTO products (name, sku, category_id, sub_category_id, brand_id, sub_brand_id, description, 
                                           barcode, stock, price_normal, price_workshop, cost_price, reorder_level)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (dialog.result['name'], sku, dialog.result['category_id'], dialog.result['sub_category_id'],
                     dialog.result['brand_id'], dialog.result['sub_brand_id'],
                     dialog.result['description'], dialog.result['barcode'],
                     dialog.result['stock'], dialog.result['price_normal'],
                     dialog.result['price_workshop'], dialog.result['cost_price'], dialog.result['reorder_level'])
                )
                
                messagebox.showinfo("Success", f"Product added successfully!\nSKU: {sku}")
                self.load_data()
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to add product: {e}")
    
    def add_category(self):
        """Add new category"""
        dialog = CategoryDialog(self.parent, "Add Category")
        if dialog.result:
            try:
                db.execute_insert(
                    "INSERT INTO categories (category_name, description) VALUES (?, ?)",
                    (dialog.result['name'], dialog.result['description'])
                )
                messagebox.showinfo("Success", "Category added successfully!")
                self.load_data()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to add category: {e}")
    
    def add_sub_category(self):
        """Add new sub category"""
        dialog = SubCategoryDialog(self.parent, "Add Sub Category")
        if dialog.result:
            try:
                db.execute_insert(
                    "INSERT INTO sub_categories (sub_category_name, category_id, description) VALUES (?, ?, ?)",
                    (dialog.result['name'], dialog.result['category_id'], dialog.result['description'])
                )
                messagebox.showinfo("Success", "Sub Category added successfully!")
                self.load_data()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to add sub category: {e}")
    
    def add_brand(self):
        """Add new brand"""
        dialog = BrandDialog(self.parent, "Add Brand")
        if dialog.result:
            try:
                db.execute_insert(
                    "INSERT INTO brands (brand_name, description) VALUES (?, ?)",
                    (dialog.result['name'], dialog.result['description'])
                )
                messagebox.showinfo("Success", "Brand added successfully!")
                self.load_data()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to add brand: {e}")
    
    def add_sub_brand(self):
        """Add new sub brand"""
        dialog = SubBrandDialog(self.parent, "Add Sub Brand")
        if dialog.result:
            try:
                db.execute_insert(
                    "INSERT INTO sub_brands (sub_brand_name, brand_id, description) VALUES (?, ?, ?)",
                    (dialog.result['name'], dialog.result['brand_id'], dialog.result['description'])
                )
                messagebox.showinfo("Success", "Sub Brand added successfully!")
                self.load_data()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to add sub brand: {e}")
    
    def edit_item(self):
        """Edit selected item"""
        selection = self.data_tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "Please select an item to edit.")
            return
        
        # Get selected item data
        item_values = self.data_tree.item(selection[0], 'values')
        
        if self.current_tab == "products":
            self.edit_product(item_values)
        elif self.current_tab == "categories":
            self.edit_category(item_values)
        elif self.current_tab == "sub_categories":
            self.edit_sub_category(item_values)
        elif self.current_tab == "brands":
            self.edit_brand(item_values)
        elif self.current_tab == "sub_brands":
            self.edit_sub_brand(item_values)
    
    def edit_product(self, item_values):
        """Edit product"""
        sku = item_values[1]
        # Get full product data
        product_query = "SELECT * FROM products WHERE sku = ?"
        product_result = db.execute_query(product_query, (sku,))
        
        if product_result:
            product = product_result[0]
            dialog = ProductDialog(self.parent, "Edit Product", product)
            if dialog.result:
                try:
                    db.execute_update(
                        """UPDATE products SET name=?, category_id=?, sub_category_id=?, brand_id=?, sub_brand_id=?, description=?,
                                             barcode=?, stock=?, price_normal=?, price_workshop=?, cost_price=?, reorder_level=?
                           WHERE sku=?""",
                        (dialog.result['name'], dialog.result['category_id'], dialog.result['sub_category_id'],
                         dialog.result['brand_id'], dialog.result['sub_brand_id'],
                         dialog.result['description'], dialog.result['barcode'],
                         dialog.result['stock'], dialog.result['price_normal'],
                         dialog.result['price_workshop'], dialog.result['cost_price'], dialog.result['reorder_level'], sku)
                    )
                    messagebox.showinfo("Success", "Product updated successfully!")
                    self.load_data()
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to update product: {e}")
    
    def edit_category(self, item_values):
        """Edit category"""
        category_id = item_values[0]
        current_name = item_values[1]
        current_desc = item_values[2] if item_values[2] != "N/A" else ""
        
        category_data = {
            'category_id': category_id,
            'category_name': current_name,
            'description': current_desc
        }
        
        dialog = CategoryDialog(self.parent, "Edit Category", category_data)
        if dialog.result:
            try:
                db.execute_update(
                    "UPDATE categories SET category_name=?, description=? WHERE category_id=?",
                    (dialog.result['name'], dialog.result['description'], category_id)
                )
                messagebox.showinfo("Success", "Category updated successfully!")
                self.load_data()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to update category: {e}")
    
    def edit_sub_category(self, item_values):
        """Edit sub category"""
        sub_category_id = item_values[0]
        current_name = item_values[1]
        current_category = item_values[2]
        current_desc = item_values[3] if item_values[3] != "N/A" else ""
        
        # Get full data
        sc_data = db.execute_query("SELECT * FROM sub_categories WHERE sub_category_id = ?", (sub_category_id,))
        if sc_data:
            sc_data = sc_data[0]
            sc_data['sub_category_name'] = current_name
            sc_data['category_name'] = current_category
            sc_data['description'] = current_desc
        else:
            sc_data = {
                'sub_category_id': sub_category_id,
                'sub_category_name': current_name,
                'category_name': current_category,
                'description': current_desc
            }
        
        dialog = SubCategoryDialog(self.parent, "Edit Sub Category", sc_data)
        if dialog.result:
            try:
                db.execute_update(
                    "UPDATE sub_categories SET sub_category_name=?, category_id=?, description=? WHERE sub_category_id=?",
                    (dialog.result['name'], dialog.result['category_id'], dialog.result['description'], sub_category_id)
                )
                messagebox.showinfo("Success", "Sub Category updated successfully!")
                self.load_data()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to update sub category: {e}")
    
    def edit_brand(self, item_values):
        """Edit brand"""
        brand_id = item_values[0]
        current_name = item_values[1]
        current_desc = item_values[2] if item_values[2] != "N/A" else ""
        
        brand_data = {
            'brand_id': brand_id,
            'brand_name': current_name,
            'description': current_desc
        }
        
        dialog = BrandDialog(self.parent, "Edit Brand", brand_data)
        if dialog.result:
            try:
                db.execute_update(
                    "UPDATE brands SET brand_name=?, description=? WHERE brand_id=?",
                    (dialog.result['name'], dialog.result['description'], brand_id)
                )
                messagebox.showinfo("Success", "Brand updated successfully!")
                self.load_data()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to update brand: {e}")
    
    def edit_sub_brand(self, item_values):
        """Edit sub brand"""
        sub_brand_id = item_values[0]
        current_name = item_values[1]
        current_brand = item_values[2]
        current_desc = item_values[3] if item_values[3] != "N/A" else ""
        
        # Get full data
        sb_data = db.execute_query("SELECT * FROM sub_brands WHERE sub_brand_id = ?", (sub_brand_id,))
        if sb_data:
            sb_data = sb_data[0]
            sb_data['sub_brand_name'] = current_name
            sb_data['brand_name'] = current_brand
            sb_data['description'] = current_desc
        else:
            sb_data = {
                'sub_brand_id': sub_brand_id,
                'sub_brand_name': current_name,
                'brand_name': current_brand,
                'description': current_desc
            }
        
        dialog = SubBrandDialog(self.parent, "Edit Sub Brand", sb_data)
        if dialog.result:
            try:
                db.execute_update(
                    "UPDATE sub_brands SET sub_brand_name=?, brand_id=?, description=? WHERE sub_brand_id=?",
                    (dialog.result['name'], dialog.result['brand_id'], dialog.result['description'], sub_brand_id)
                )
                messagebox.showinfo("Success", "Sub Brand updated successfully!")
                self.load_data()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to update sub brand: {e}")
    
    def delete_item(self):
        """Delete selected item"""
        selection = self.data_tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "Please select an item to delete.")
            return
        
        if not messagebox.askyesno("Confirm Delete", "Are you sure you want to delete this item?"):
            return
        
        item_values = self.data_tree.item(selection[0], 'values')
        
        try:
            if self.current_tab == "products":
                sku = item_values[1]
                db.execute_update("UPDATE products SET is_active=0 WHERE sku=?", (sku,))
            elif self.current_tab == "categories":
                category_id = item_values[0]
                db.execute_update("DELETE FROM categories WHERE category_id=?", (category_id,))
            elif self.current_tab == "sub_categories":
                sub_category_id = item_values[0]
                db.execute_update("DELETE FROM sub_categories WHERE sub_category_id=?", (sub_category_id,))
            elif self.current_tab == "brands":
                brand_id = item_values[0]
                db.execute_update("DELETE FROM brands WHERE brand_id=?", (brand_id,))
            elif self.current_tab == "sub_brands":
                sub_brand_id = item_values[0]
                db.execute_update("DELETE FROM sub_brands WHERE sub_brand_id=?", (sub_brand_id,))
            
            messagebox.showinfo("Success", "Item deleted successfully!")
            self.load_data()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete item: {e}")
    
    def export_to_excel(self):
        """Export inventory to Excel"""
        if self.current_tab != "products":
            messagebox.showinfo("Info", "Excel export is only available for Products tab.")
            return
        
        ExportManager.export_inventory_to_excel(
            filter_category=self.category_filter_var.get(),
            filter_brand=self.brand_filter_var.get(),
            filter_stock=self.stock_filter_var.get()
        )
    
    def export_to_pdf(self):
        """Export inventory to PDF"""
        if self.current_tab != "products":
            messagebox.showinfo("Info", "PDF export is only available for Products tab.")
            return
        
        ExportManager.export_inventory_to_pdf(
            filter_category=self.category_filter_var.get(),
            filter_brand=self.brand_filter_var.get(),
            filter_stock=self.stock_filter_var.get()
        )
    
    def handle_export(self, choice):
        """Handle export dropdown selection"""
        if "Excel" in choice:
            self.export_to_excel()
        elif "PDF" in choice:
            self.export_to_pdf()
        self.export_menu.set("📥 Export")

    def import_from_excel(self):
        """Import products from Excel using smart auto-detection.
        
        Workflow:
          1. User selects an Excel file
          2. All sheets are scanned automatically (no manual column mapping)
          3. A preview dialog shows what will be imported
          4. On confirmation a progress dialog runs the actual import
          5. Inventory table is refreshed
        
        Duplicate handling: products whose name already exists in the same
        category are silently skipped.
        """
        if self.current_tab != "products":
            messagebox.showinfo("Info", "Excel import is only available for Products tab.")
            return

        from utils.excel_importer import SimpleExcelImporter

        importer = SimpleExcelImporter(self.parent)

        # Step 1: browse
        file_path = importer.browse_file()
        if not file_path:
            return  # user cancelled

        # Step 2: scan & preview
        preview_data = importer.extract_preview_data(file_path)
        if not preview_data:
            messagebox.showwarning(
                "No Data Found",
                "No importable product sheets were found in this file.\n\n"
                "Make sure your Excel has columns like:\n"
                "  • Items / Bikes names / Product Name\n"
                "  • Price / Wholesale Price\n"
                "  • Qty in Hand / Stock"
            )
            return

        # Step 3: show preview dialog
        proceed = self._show_preview_dialog(preview_data)
        if not proceed:
            return

        # Step 4: import with progress
        self._do_import_with_progress(importer, file_path)
    
    def _show_preview_dialog(self, preview_data):
        """Show preview of data before import - smaller and responsive"""
        # Calculate responsive dimensions
        parent_width = self.parent.winfo_width()
        parent_height = self.parent.winfo_height()
        if parent_width < 100:
            parent_width = 1200
        if parent_height < 100:
            parent_height = 800
        
        dialog_width = int(min(parent_width * 0.7, 800))
        dialog_height = int(min(parent_height * 0.7, 550))
        
        preview_dialog = ctk.CTkToplevel(self.parent)
        preview_dialog.title("Preview Excel Data")
        preview_dialog.geometry(f"{dialog_width}x{dialog_height}")
        preview_dialog.transient(self.parent)
        preview_dialog.grab_set()
        
        # Center dialog with bounds checking
        preview_dialog.update_idletasks()
        x = self.parent.winfo_x() + (self.parent.winfo_width() // 2) - (dialog_width // 2)
        y = self.parent.winfo_y() + (self.parent.winfo_height() // 2) - (dialog_height // 2)
        x = max(0, min(x, preview_dialog.winfo_screenwidth() - dialog_width))
        y = max(0, min(y, preview_dialog.winfo_screenheight() - dialog_height))
        preview_dialog.geometry(f"{dialog_width}x{dialog_height}+{x}+{y}")
        
        # Header
        header = ctk.CTkFrame(preview_dialog, corner_radius=10, fg_color=("#3B82F6", "#2563EB"))
        header.pack(fill="x", padx=15, pady=(15, 8))
        ctk.CTkLabel(header, text="📋 Preview Excel Data", 
                    font=ctk.CTkFont(size=16, weight="bold"), text_color="white").pack(pady=10)
        
        # Summary
        total_sheets = len(preview_data)
        total_rows = sum(sheet['total_rows'] for sheet in preview_data)
        
        summary_frame = ctk.CTkFrame(preview_dialog)
        summary_frame.pack(fill="x", padx=15, pady=(0, 8))
        
        ctk.CTkLabel(summary_frame, 
                    text=f"Found {total_sheets} sheets with approximately {total_rows} products",
                    font=ctk.CTkFont(size=12, weight="bold")).pack(pady=8)
        
        # Preview tree
        tree_frame = ctk.CTkFrame(preview_dialog)
        tree_frame.pack(fill="both", expand=True, padx=15, pady=8)
        
        # Calculate columns based on dialog width
        col_width = max(100, (dialog_width - 50) // 6)
        name_width = max(200, (dialog_width - 50) - (col_width * 5))
        
        tree = ttk.Treeview(tree_frame, columns=("Sheet", "Product Name", "Brand", "Cost", "Price", "Stock"), 
                           show="headings", height=12)
        tree.pack(side="left", fill="both", expand=True)
        
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        scrollbar.pack(side="right", fill="y")
        tree.configure(yscrollcommand=scrollbar.set)
        
        # Configure columns
        for col in ("Sheet", "Product Name", "Brand", "Cost", "Price", "Stock"):
            tree.heading(col, text=col)
            tree.column(col, width=col_width if col != "Product Name" else name_width)
        
        # Insert preview data
        for sheet_data in preview_data:
            for product in sheet_data['sample_products']:
                # Ensure name is a string
                name = str(product['name']) if product['name'] else ""
                tree.insert("", "end", values=(
                    sheet_data['sheet_name'],
                    name[:40],
                    product['brand'],
                    f"₹{product['cost']:.0f}",
                    f"₹{product['price']:.0f}",
                    product['stock']
                ))
        
        # Buttons - responsive sizing
        button_frame = ctk.CTkFrame(preview_dialog, fg_color="transparent")
        button_frame.pack(fill="x", padx=15, pady=(8, 15))
        
        result = {"proceed": False}
        
        def on_import():
            result["proceed"] = True
            preview_dialog.destroy()
        
        def on_cancel():
            result["proceed"] = False
            preview_dialog.destroy()
        
        # Responsive button sizing
        btn_width = max(120, min(180, dialog_width // 5))
        ctk.CTkButton(button_frame, text="✓ Import This Data", command=on_import,
                     fg_color=("#10B981", "#059669"), hover_color=("#059669", "#047857"),
                     width=btn_width, height=40, font=ctk.CTkFont(size=12, weight="bold")).pack(side="left", padx=(0, 8))
        
        ctk.CTkButton(button_frame, text="✗ Cancel", command=on_cancel,
                     fg_color=("#EF4444", "#B91C1C"), hover_color=("#DC2626", "#991B1B"),
                     width=int(btn_width * 0.7), height=40, font=ctk.CTkFont(size=12)).pack(side="left")
        
        self.parent.wait_window(preview_dialog)
        return result["proceed"]
    
    def _do_import_with_progress(self, importer, file_path):
        """Show import progress and execute import in background thread"""
        parent_width = self.parent.winfo_width()
        parent_height = self.parent.winfo_height()
        if parent_width < 100:
            parent_width = 1200
        if parent_height < 100:
            parent_height = 800
        
        dialog_width = int(min(parent_width * 0.45, 450))
        dialog_height = int(min(parent_height * 0.5, 350))
        
        progress_dialog = ctk.CTkToplevel(self.parent)
        progress_dialog.title("Importing...")
        progress_dialog.geometry(f"{dialog_width}x{dialog_height}")
        progress_dialog.transient(self.parent)
        progress_dialog.grab_set()
        
        # Center dialog
        progress_dialog.update_idletasks()
        x = self.parent.winfo_x() + (self.parent.winfo_width() // 2) - (dialog_width // 2)
        y = self.parent.winfo_y() + (self.parent.winfo_height() // 2) - (dialog_height // 2)
        x = max(0, min(x, progress_dialog.winfo_screenwidth() - dialog_width))
        y = max(0, min(y, progress_dialog.winfo_screenheight() - dialog_height))
        progress_dialog.geometry(f"{dialog_width}x{dialog_height}+{x}+{y}")
        
        # Header
        header = ctk.CTkFrame(progress_dialog, corner_radius=10, fg_color=("#10B981", "#059669"))
        header.pack(fill="x", padx=15, pady=(12, 8))
        ctk.CTkLabel(header, text="📥 Importing Data",
                    font=ctk.CTkFont(size=14, weight="bold"), text_color="white").pack(pady=10)
        
        # Progress bar
        progress_bar = ctk.CTkProgressBar(progress_dialog)
        progress_bar.pack(fill="x", padx=15, pady=(0, 5))
        progress_bar.set(0)
        
        # Progress text
        progress_text_height = int(dialog_height * 0.35)
        progress_text = ctk.CTkTextbox(progress_dialog, height=progress_text_height, font=ctk.CTkFont(size=10))
        progress_text.pack(fill="both", expand=True, padx=15, pady=8)
        progress_text.insert("1.0", "Starting import...\n\n")
        
        # Status label
        status_label = ctk.CTkLabel(progress_dialog, text="Importing...",
                                   font=ctk.CTkFont(size=11, weight="bold"))
        status_label.pack(pady=(0, 8))
        
        # Close button (disabled during import)
        close_btn = ctk.CTkButton(progress_dialog, text="Close",
                                 command=progress_dialog.destroy,
                                 state="disabled", width=100, height=35)
        close_btn.pack(pady=(0, 12))
        
        # --- Thread-safe UI update helpers ---
        def _append_text(msg):
            progress_text.insert("end", f"{msg}\n")
            progress_text.see("end")
        
        def _on_complete(total_imported):
            progress_bar.set(1.0)
            _append_text("\n" + "=" * 40 + "\n")
            _append_text("✓ IMPORT COMPLETE!\n\n")

            stats = getattr(importer, "stats", {}) or {}

            _append_text(f"Sheets processed:    {stats.get('sheets_processed', 0)}\n")
            _append_text(f"Products imported:   {total_imported}\n")
            _append_text(f"Duplicates skipped:  {stats.get('products_skipped', 0)}\n")
            _append_text(f"Categories created:  {stats.get('categories_created', 0)}\n")
            _append_text(f"Brands created:      {stats.get('brands_created', 0)}\n")

            errors = stats.get('errors', []) or []
            if errors:
                _append_text(f"\nErrors: {len(errors)}\n")
                for error in errors[:5]:
                    _append_text(f"  - {error}\n")

            status_label.configure(text="Import completed!", text_color=("#10B981", "#34D399"))
            close_btn.configure(state="normal", text="Close & Refresh")
            self.load_data()
        
        def _on_error(error_msg):
            progress_bar.set(0)
            _append_text(f"\n✗ ERROR: {error_msg}\n")
            status_label.configure(text="Import failed!", text_color=("#EF4444", "#F87171"))
            close_btn.configure(state="normal", text="Close")
        
        # --- Background import thread ---
        def thread_target():
            def callback(msg):
                progress_dialog.after(0, lambda: _append_text(msg))
            
            try:
                total_imported = importer.import_file(file_path, progress_callback=callback)
                progress_dialog.after(0, lambda: _on_complete(total_imported))
            except Exception as e:
                progress_dialog.after(0, lambda: _on_error(str(e)))
        
        threading.Thread(target=thread_target, daemon=True).start()


class ExcelImportDialog:
    """Dialog for importing products from Excel with column mapping"""
    def __init__(self, parent, title):
        self.parent = parent
        self.result = None
        self.excel_data = None
        self.excel_columns = []
        self.preview_data = []
        
        # Database field mappings
        self.db_fields = {
            'name': {'label': 'Product Name*', 'required': True, 'type': 'text'},
            'sku': {'label': 'SKU*', 'required': True, 'type': 'text'},
            'category': {'label': 'Category', 'required': False, 'type': 'text'},
            'sub_category': {'label': 'Sub Category', 'required': False, 'type': 'text'},
            'brand': {'label': 'Brand', 'required': False, 'type': 'text'},
            'sub_brand': {'label': 'Sub Brand', 'required': False, 'type': 'text'},
            'description': {'label': 'Description', 'required': False, 'type': 'text'},
            'stock': {'label': 'Stock Quantity', 'required': False, 'type': 'number', 'default': 0},
            'cost_price': {'label': 'Cost Price*', 'required': True, 'type': 'number'},
            'price_normal': {'label': 'Normal Price*', 'required': True, 'type': 'number'},
            'price_workshop': {'label': 'Workshop Price*', 'required': True, 'type': 'number'},
            'reorder_level': {'label': 'Reorder Level', 'required': False, 'type': 'number', 'default': 10}
        }
        
        self.field_mappings = {}  # Will hold db_field -> excel_column mappings
        
        # Calculate responsive dimensions
        self._calculate_responsive_dimensions()
        
        # Create dialog
        self.dialog = ctk.CTkToplevel(parent)
        self.dialog.title(title)
        self.dialog.geometry(f"{self.dialog_width}x{self.dialog_height}")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # Center dialog on parent window
        self._center_dialog()
        
        # Create UI
        self.create_ui()
        
        # Wait for dialog to close
        self.dialog.wait_window()
    
    def _calculate_responsive_dimensions(self):
        """Calculate responsive window dimensions based on parent window"""
        parent_width = self.parent.winfo_width()
        parent_height = self.parent.winfo_height()
        
        # Handle case where parent window might not be fully rendered
        if parent_width < 100 or parent_height < 100:
            parent_width = 1200
            parent_height = 800
        
        # Calculate dialog size as percentage of parent (max 85% of parent)
        dialog_width = int(min(parent_width * 0.85, 1100))
        dialog_height = int(min(parent_height * 0.85, 800))
        
        # Set minimum dimensions
        self.dialog_width = max(dialog_width, 800)
        self.dialog_height = max(dialog_height, 650)
    
    def _center_dialog(self):
        """Center the dialog on the parent window"""
        self.dialog.update_idletasks()
        self.parent.update_idletasks()
        
        parent_x = self.parent.winfo_rootx()
        parent_y = self.parent.winfo_rooty()
        parent_width = self.parent.winfo_width()
        parent_height = self.parent.winfo_height()
        
        # Handle case where parent window might not be fully rendered
        if parent_width < 100:
            parent_width = 1200
        if parent_height < 100:
            parent_height = 800
        
        x = parent_x + (parent_width // 2) - (self.dialog_width // 2)
        y = parent_y + (parent_height // 2) - (self.dialog_height // 2)
        
        # Ensure dialog stays within screen bounds
        screen_width = self.dialog.winfo_screenwidth()
        screen_height = self.dialog.winfo_screenheight()
        
        x = max(0, min(x, screen_width - self.dialog_width))
        y = max(0, min(y, screen_height - self.dialog_height))
        
        self.dialog.geometry(f"{self.dialog_width}x{self.dialog_height}+{x}+{y}")
    
    def create_ui(self):
        """Create the import dialog UI"""
        # Main container
        main_frame = ctk.CTkScrollableFrame(self.dialog)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Header
        header = ctk.CTkFrame(main_frame, corner_radius=10, fg_color="#10B981")
        header.pack(fill="x", pady=(0, 20))
        ctk.CTkLabel(header, text="📥 Import Products from Excel", 
                     font=ctk.CTkFont(size=20, weight="bold"), text_color="white").pack(pady=15)
        
        # Step 1: File Selection
        file_frame = ctk.CTkFrame(main_frame)
        file_frame.pack(fill="x", pady=(0, 20))
        
        ctk.CTkLabel(file_frame, text="Step 1: Select Excel File", 
                     font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", padx=15, pady=(15, 10))
        
        file_select_frame = ctk.CTkFrame(file_frame, fg_color="transparent")
        file_select_frame.pack(fill="x", padx=15, pady=(0, 15))
        
        self.file_path_var = tk.StringVar()
        self.file_entry = ctk.CTkEntry(file_select_frame, textvariable=self.file_path_var, 
                                       placeholder_text="Select an Excel file...", width=500)
        self.file_entry.pack(side="left", padx=(0, 10))
        
        ctk.CTkButton(file_select_frame, text="Browse...", command=self.browse_file,
                     fg_color="#3B82F6", hover_color="#2563EB").pack(side="left")
        
        # Step 2: Column Mapping
        self.mapping_frame = ctk.CTkFrame(main_frame)
        self.mapping_frame.pack(fill="x", pady=(0, 20))
        
        ctk.CTkLabel(self.mapping_frame, text="Step 2: Map Excel Columns to Database Fields", 
                     font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", padx=15, pady=(15, 10))
        
        ctk.CTkLabel(self.mapping_frame, text="Select which Excel column corresponds to each database field.", 
                     font=ctk.CTkFont(size=11), text_color="gray").pack(anchor="w", padx=15, pady=(0, 10))
        
        # Create mapping dropdowns
        self.mapping_widgets = {}
        mapping_grid = ctk.CTkFrame(self.mapping_frame, fg_color="transparent")
        mapping_grid.pack(fill="x", padx=15, pady=(0, 15))
        
        row = 0
        col = 0
        for field_key, field_info in self.db_fields.items():
            field_frame = ctk.CTkFrame(mapping_grid, fg_color="transparent")
            field_frame.grid(row=row, column=col, padx=10, pady=5, sticky="w")
            
            label_text = field_info['label']
            ctk.CTkLabel(field_frame, text=label_text, font=ctk.CTkFont(size=11, weight="bold")).pack(anchor="w")
            
            dropdown = ctk.CTkOptionMenu(
                field_frame, 
                values=["-- Not Mapped --"], 
                width=200,
                command=lambda x, fk=field_key: self.on_mapping_changed(fk, x)
            )
            dropdown.pack(anchor="w", pady=(2, 0))
            
            self.mapping_widgets[field_key] = dropdown
            
            col += 1
            if col > 2:
                col = 0
                row += 1
        
        # Step 3: Preview
        preview_frame = ctk.CTkFrame(main_frame)
        preview_frame.pack(fill="both", expand=True, pady=(0, 20))
        
        ctk.CTkLabel(preview_frame, text="Step 3: Preview Data", 
                     font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", padx=15, pady=(15, 10))
        
        # Preview treeview
        self.preview_tree = ttk.Treeview(preview_frame, show="headings", height=8)
        self.preview_tree.pack(side="left", fill="both", expand=True, padx=(15, 5), pady=(0, 15))
        
        scrollbar = ttk.Scrollbar(preview_frame, orient="vertical", command=self.preview_tree.yview)
        scrollbar.pack(side="right", fill="y", padx=(0, 15), pady=(0, 15))
        self.preview_tree.configure(yscrollcommand=scrollbar.set)
        
        # Buttons
        button_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        button_frame.pack(fill="x", pady=(0, 10))
        
        self.import_btn = ctk.CTkButton(button_frame, text="Import Products", command=self.import_data,
                                       fg_color="#10B981", hover_color="#059669",
                                       font=ctk.CTkFont(size=13, weight="bold"),
                                       width=150, height=40, state="disabled")
        self.import_btn.pack(side="left", padx=(0, 10))
        
        ctk.CTkButton(button_frame, text="Cancel", command=self.cancel,
                     fg_color="#6B7280", hover_color="#4B5563",
                     width=100, height=40).pack(side="left")
        
        # Status label
        self.status_label = ctk.CTkLabel(main_frame, text="", font=ctk.CTkFont(size=11))
        self.status_label.pack(anchor="w", pady=(10, 0))
    
    def browse_file(self):
        """Browse for Excel file"""
        from tkinter import filedialog
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            self.file_path_var.set(file_path)
            self.load_excel_file(file_path)
    
    def load_excel_file(self, file_path):
        """Load and parse Excel file with intelligent header detection"""
        try:
            import openpyxl
            
            # Load workbook
            wb = openpyxl.load_workbook(file_path)
            
            # Check if multiple sheets exist
            sheet_names = wb.sheetnames
            if len(sheet_names) > 1:
                # Show sheet selector with ALL sheets
                selected_sheet = self.show_sheet_selector(wb, sheet_names)
                if not selected_sheet:
                    return  # User cancelled
                
                sheet = wb[selected_sheet]
            else:
                sheet = wb.active
            
            # Show first few rows to help identify header row (show first 15 columns for more data)
            preview_rows = []
            for row_num in range(1, min(10, sheet.max_row + 1)):
                row = sheet[row_num]
                row_data = [str(cell.value)[:25] if cell.value else "" for cell in row[:15]]  # Show 15 columns
                preview_rows.append(f"Row {row_num}: {' | '.join(row_data)}")
            
            # Calculate responsive dimensions
            parent_width = self.dialog.winfo_width()
            parent_height = self.dialog.winfo_height()
            if parent_width < 100:
                parent_width = 1000
            if parent_height < 100:
                parent_height = 700
            
            dialog_width = int(min(parent_width * 0.65, 800))
            dialog_height = int(min(parent_height * 0.75, 550))
            
            # Ask user to select header row
            header_dialog = tk.Toplevel(self.dialog)
            header_dialog.title("Select Header Row")
            header_dialog.geometry(f"{dialog_width}x{dialog_height}")
            header_dialog.transient(self.dialog)
            header_dialog.grab_set()
            header_dialog.configure(bg="#2D2D2D")
            
            # Center dialog with bounds checking
            header_dialog.update_idletasks()
            x = self.dialog.winfo_x() + (self.dialog.winfo_width() // 2) - (dialog_width // 2)
            y = self.dialog.winfo_y() + (self.dialog.winfo_height() // 2) - (dialog_height // 2)
            x = max(0, min(x, header_dialog.winfo_screenwidth() - dialog_width))
            y = max(0, min(y, header_dialog.winfo_screenheight() - dialog_height))
            header_dialog.geometry(f"{dialog_width}x{dialog_height}+{x}+{y}")
            
            ctk.CTkLabel(header_dialog, text="Preview of Excel file (first 15 columns):", 
                        font=ctk.CTkFont(size=12, weight="bold")).pack(pady=10)
            
            # Show preview with scrolling
            preview_frame = ctk.CTkFrame(header_dialog, fg_color="#1E1E1E")
            preview_frame.pack(fill="both", expand=True, padx=10, pady=5)
            
            preview_text = tk.Text(preview_frame, height=12, width=90, font=("Courier", 9), 
                                  bg="#1E1E1E", fg="white", wrap="none")
            scrollbar_y = ttk.Scrollbar(preview_frame, orient="vertical", command=preview_text.yview)
            scrollbar_x = ttk.Scrollbar(preview_frame, orient="horizontal", command=preview_text.xview)
            preview_text.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
            
            preview_text.pack(side="left", fill="both", expand=True)
            scrollbar_y.pack(side="right", fill="y")
            scrollbar_x.pack(side="bottom", fill="x")
            
            preview_text.insert("1.0", "\n".join(preview_rows))
            preview_text.configure(state="disabled")
            
            ctk.CTkLabel(header_dialog, text="Select which row contains the column headers:", 
                        font=ctk.CTkFont(size=12)).pack(pady=10)
            
            row_var = tk.IntVar(value=1)
            row_frame = ctk.CTkFrame(header_dialog)
            row_frame.pack(pady=10)
            
            for i in range(1, min(6, sheet.max_row + 1)):
                ctk.CTkRadioButton(row_frame, text=f"Row {i}", variable=row_var, value=i).pack(anchor="w", pady=2)
            
            def on_confirm():
                header_dialog.destroy()
                self._process_excel_with_header_row(sheet, row_var.get())
            
            ctk.CTkButton(header_dialog, text="Confirm", command=on_confirm, 
                         fg_color="#10B981", width=100).pack(pady=20)
            
            self.dialog.wait_window(header_dialog)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load Excel file:\n{e}")
            self.status_label.configure(text="✗ Failed to load file", text_color="#EF4444")
    
    def _process_excel_with_header_row(self, sheet, header_row_idx):
        """Process Excel file with specified header row"""
        try:
            # Get headers from the specified header row
            raw_headers = []
            for i, cell in enumerate(sheet[header_row_idx]):
                if cell.value:
                    raw_headers.append(str(cell.value).strip())
                else:
                    raw_headers.append(f"Column_{i+1}")
            
            pass
            
            # Data starts from the row after the header
            data_start_row = header_row_idx + 1
            
            # Get first data row for samples
            first_row = []
            for row in sheet.iter_rows(min_row=data_start_row, max_row=data_start_row, values_only=True):
                first_row = [str(val) if val is not None else "" for val in row]
                break
            
            # Create column options with samples
            self.excel_columns = raw_headers
            self.column_samples = {}
            
            column_options = ["-- Not Mapped --"]
            for i, header in enumerate(raw_headers):
                sample = first_row[i] if i < len(first_row) else ""
                # Truncate sample if too long
                sample_display = sample[:20] + "..." if len(sample) > 20 else sample
                if sample_display:
                    option_text = f"{header}  |  Ex: {sample_display}"
                else:
                    option_text = header
                self.column_samples[header] = option_text
                column_options.append(option_text)
            
            pass
            
            # Get preview data (first 5 rows)
            self.preview_data = []
            for row in sheet.iter_rows(min_row=data_start_row, max_row=data_start_row + 4, values_only=True):
                self.preview_data.append([str(val) if val is not None else "" for val in row])
            
            # Get all data for import
            self.excel_data = []
            for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
                self.excel_data.append([str(val) if val is not None else "" for val in row])
            
            # Store raw headers for mapping
            self.raw_column_options = ["-- Not Mapped --"] + raw_headers
            
            # Update dropdowns with column options showing samples
            pass
            for field_key, dropdown in self.mapping_widgets.items():
                try:
                    dropdown.configure(values=column_options)
                    pass
                except Exception as dropdown_error:
                    pass
            
            # Try to auto-map columns based on name matching
            self.auto_map_columns()
            
            # Update preview
            self.update_preview()
            
            # Enable import button
            self.import_btn.configure(state="normal")
            
            self.status_label.configure(text="✓ Loaded " + str(len(self.excel_data)) + " rows from Excel file", 
                                       text_color="#10B981")
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            messagebox.showerror("Error", f"Failed to load Excel file:\n{str(e)}\n\nSee console for details.")
            self.status_label.configure(text="✗ Failed to load file", text_color="#EF4444")
    
    def auto_map_columns(self):
        """Automatically map columns based on name similarity"""
        excel_cols_lower = {col.lower(): col for col in self.excel_columns}
        
        for field_key, field_info in self.db_fields.items():
            # Check for exact match
            if field_key.lower() in excel_cols_lower:
                self.mapping_widgets[field_key].set(excel_cols_lower[field_key.lower()])
                continue
            
            # Check for label match
            label_words = field_info['label'].lower().replace('*', '').strip().split()
            for col in self.excel_columns:
                col_lower = col.lower()
                # Check if any significant word matches
                for word in label_words:
                    if len(word) > 2 and word in col_lower:
                        self.mapping_widgets[field_key].set(col)
                        break
                else:
                    continue
                break
    
    def update_preview(self):
        """Update preview treeview to show mapped data preview"""
        # Clear existing columns and items
        for item in self.preview_tree.get_children():
            self.preview_tree.delete(item)
        self.preview_tree['columns'] = []
        
        if not self.excel_columns or not self.preview_data:
            return
        
        # Build columns: Show Excel column name + mapped field name
        display_columns = []
        
        # Get all Excel columns that are mapped
        for field_key, dropdown in self.mapping_widgets.items():
            mapped_col = dropdown.get()
            if mapped_col != "-- Not Mapped --" and mapped_col in self.excel_columns:
                col_idx = self.excel_columns.index(mapped_col)
                col_id = f"col_{col_idx}"
                # Show both Excel column name and field name
                header_text = f"{mapped_col}\n({self.db_fields[field_key]['label']})"
                display_columns.append((col_id, header_text, col_idx))
        
        if not display_columns:
            # No mappings yet, show first few Excel columns
            for i, col_name in enumerate(self.excel_columns[:5]):
                col_id = f"col_{i}"
                display_columns.append((col_id, col_name, i))
        
        # Configure columns
        self.preview_tree['columns'] = [c[0] for c in display_columns]
        
        for col_id, header_text, col_idx in display_columns:
            # Truncate long headers
            display_header = header_text[:25] + "..." if len(header_text) > 25 else header_text
            self.preview_tree.heading(col_id, text=display_header, anchor="center")
            self.preview_tree.column(col_id, width=120, anchor="center")
        
        # Populate data rows (first 5 rows)
        for row_data in self.preview_data[:5]:
            values = []
            for col_id, header_text, col_idx in display_columns:
                if col_idx < len(row_data):
                    val = str(row_data[col_idx])
                    # Truncate long values
                    values.append(val[:25] + "..." if len(val) > 25 else val)
                else:
                    values.append("")
            
            self.preview_tree.insert("", "end", values=values)
    
    def import_data(self):
        """Import data into database with preview confirmation"""
        # Validate required fields
        for field_key, field_info in self.db_fields.items():
            if field_info['required']:
                dropdown = self.mapping_widgets[field_key]
                selected = dropdown.get()
                if selected == "-- Not Mapped --":
                    messagebox.showerror("Validation Error", 
                                        f"Required field '{field_info['label']}' is not mapped!")
                    return
        
        # Get mappings
        mappings = {}
        for field_key, dropdown in self.mapping_widgets.items():
            selected_display = dropdown.get()
            if selected_display != "-- Not Mapped --":
                raw_col_name = selected_display.split("  |  ")[0] if "  |  " in selected_display else selected_display
                if raw_col_name in self.excel_columns:
                    mappings[field_key] = self.excel_columns.index(raw_col_name)
        
        # Analyze data first (don't import yet)
        preview_data = self._analyze_import_data(mappings)
        
        if preview_data['total_products'] == 0:
            messagebox.showerror("Import Error", "No valid products found to import!")
            return
        
        # Show comprehensive preview dialog
        if not self._show_import_preview_dialog(preview_data):
            return  # User cancelled
        
        # Now actually import the data
        imported_count = self._do_import(mappings, preview_data)
        
        self.result = {'imported_count': imported_count}
        self.dialog.destroy()
    
    def _analyze_import_data(self, mappings):
        """Analyze Excel data and return preview information"""
        preview = {
            'total_products': 0,
            'new_categories': set(),
            'new_brands': set(),
            'existing_skus': [],
            'sample_products': [],
            'errors': []
        }
        
        for row_idx, row_data in enumerate(self.excel_data):
            try:
                # Extract values
                values = {}
                for field_key, col_idx in mappings.items():
                    val = row_data[col_idx] if col_idx < len(row_data) else ""
                    field_info = self.db_fields[field_key]
                    if field_info['type'] == 'number':
                        try:
                            val = float(val) if val else (field_info.get('default', 0))
                        except:
                            val = field_info.get('default', 0)
                    values[field_key] = val
                
                # Skip if missing required fields
                has_required = True
                for field_key, field_info in self.db_fields.items():
                    if field_info['required'] and not values.get(field_key):
                        has_required = False
                        break
                
                if not has_required:
                    continue
                
                # Track categories and brands
                category = values.get('category', '').strip()
                brand = values.get('brand', '').strip()
                
                if category:
                    # Check if category exists
                    existing = db.execute_query("SELECT category_id FROM categories WHERE LOWER(category_name) = LOWER(?)", (category,))
                    if not existing:
                        preview['new_categories'].add(category)
                
                if brand:
                    existing = db.execute_query("SELECT brand_id FROM brands WHERE LOWER(brand_name) = LOWER(?)", (brand,))
                    if not existing:
                        preview['new_brands'].add(brand)
                
                # Check for existing SKU
                sku = values.get('sku', '').strip()
                if sku:
                    existing = db.execute_query("SELECT product_id FROM products WHERE sku = ?", (sku,))
                    if existing:
                        preview['existing_skus'].append(sku)
                
                # Add to sample products (first 10)
                if len(preview['sample_products']) < 10:
                    preview['sample_products'].append({
                        'name': values.get('name', ''),
                        'category': category or 'N/A',
                        'brand': brand or 'N/A',
                        'stock': values.get('stock', 0),
                        'cost_price': values.get('cost_price', 0),
                        'price_normal': values.get('price_normal', 0)
                    })
                
                preview['total_products'] += 1
                
            except Exception as e:
                preview['errors'].append(f"Row {row_idx + 2}: {str(e)}")
        
        return preview
    
    def _show_import_preview_dialog(self, preview_data):
        """Show preview dialog with import summary"""
        # Calculate responsive dimensions
        parent_width = self.dialog.winfo_width()
        parent_height = self.dialog.winfo_height()
        if parent_width < 100:
            parent_width = 1000
        if parent_height < 100:
            parent_height = 700
        
        dialog_width = int(min(parent_width * 0.85, 1000))
        dialog_height = int(min(parent_height * 0.9, 750))
        
        preview_dialog = ctk.CTkToplevel(self.dialog)
        preview_dialog.title("Import Preview - Review Before Importing")
        preview_dialog.geometry(f"{dialog_width}x{dialog_height}")
        preview_dialog.transient(self.dialog)
        preview_dialog.grab_set()
        
        # Center dialog with bounds checking
        preview_dialog.update_idletasks()
        x = self.dialog.winfo_x() + (self.dialog.winfo_width() // 2) - (dialog_width // 2)
        y = self.dialog.winfo_y() + (self.dialog.winfo_height() // 2) - (dialog_height // 2)
        x = max(0, min(x, preview_dialog.winfo_screenwidth() - dialog_width))
        y = max(0, min(y, preview_dialog.winfo_screenheight() - dialog_height))
        preview_dialog.geometry(f"{dialog_width}x{dialog_height}+{x}+{y}")
        
        # Main container
        main_frame = ctk.CTkScrollableFrame(preview_dialog)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Header
        header = ctk.CTkFrame(main_frame, corner_radius=10, fg_color="#10B981")
        header.pack(fill="x", pady=(0, 20))
        ctk.CTkLabel(header, text="📋 Import Preview", 
                    font=ctk.CTkFont(size=20, weight="bold"), text_color="white").pack(pady=15)
        
        # Summary Section
        summary_frame = ctk.CTkFrame(main_frame)
        summary_frame.pack(fill="x", pady=(0, 15))
        
        ctk.CTkLabel(summary_frame, text="📊 Import Summary", 
                    font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", padx=15, pady=(15, 10))
        
        summary_grid = ctk.CTkFrame(summary_frame, fg_color="transparent")
        summary_grid.pack(fill="x", padx=15, pady=(0, 15))
        
        # Stats boxes
        stats = [
            ("📦 Products", str(preview_data['total_products']), "#3B82F6"),
            ("🏷️ New Categories", str(len(preview_data['new_categories'])), "#8B5CF6"),
            ("🏭 New Brands", str(len(preview_data['new_brands'])), "#F59E0B"),
        ]
        
        for i, (label, value, color) in enumerate(stats):
            stat_box = ctk.CTkFrame(summary_grid, fg_color=color, corner_radius=8)
            stat_box.grid(row=0, column=i, padx=5, pady=5, sticky="ew")
            stat_box.grid_columnconfigure(0, weight=1)
            ctk.CTkLabel(stat_box, text=label, font=ctk.CTkFont(size=11), text_color="white").pack(pady=(8, 2))
            ctk.CTkLabel(stat_box, text=value, font=ctk.CTkFont(size=20, weight="bold"), text_color="white").pack(pady=(0, 8))
        
        # New Categories Section
        if preview_data['new_categories']:
            cat_frame = ctk.CTkFrame(main_frame)
            cat_frame.pack(fill="x", pady=(0, 15))
            ctk.CTkLabel(cat_frame, text="🏷️ New Categories to Create", 
                        font=ctk.CTkFont(size=13, weight="bold")).pack(anchor="w", padx=15, pady=(10, 5))
            cat_text = ", ".join(sorted(list(preview_data['new_categories']))[:20])
            if len(preview_data['new_categories']) > 20:
                cat_text += f"... and {len(preview_data['new_categories']) - 20} more"
            ctk.CTkLabel(cat_frame, text=cat_text, font=ctk.CTkFont(size=11), 
                        wraplength=700).pack(anchor="w", padx=15, pady=(0, 10))
        
        # New Brands Section
        if preview_data['new_brands']:
            brand_frame = ctk.CTkFrame(main_frame)
            brand_frame.pack(fill="x", pady=(0, 15))
            ctk.CTkLabel(brand_frame, text="🏭 New Brands to Create", 
                        font=ctk.CTkFont(size=13, weight="bold")).pack(anchor="w", padx=15, pady=(10, 5))
            brand_text = ", ".join(sorted(list(preview_data['new_brands']))[:20])
            if len(preview_data['new_brands']) > 20:
                brand_text += f"... and {len(preview_data['new_brands']) - 20} more"
            ctk.CTkLabel(brand_frame, text=brand_text, font=ctk.CTkFont(size=11), 
                        wraplength=700).pack(anchor="w", padx=15, pady=(0, 10))
        
        # Existing SKUs Section
        if preview_data['existing_skus']:
            existing_frame = ctk.CTkFrame(main_frame, fg_color="#FEF3C7")
            existing_frame.pack(fill="x", pady=(0, 15))
            ctk.CTkLabel(existing_frame, text="⚠️ Existing SKUs (Will be updated)", 
                        font=ctk.CTkFont(size=13, weight="bold"), text_color="#92400E").pack(anchor="w", padx=15, pady=(10, 5))
            sku_text = ", ".join(preview_data['existing_skus'][:10])
            if len(preview_data['existing_skus']) > 10:
                sku_text += f"... and {len(preview_data['existing_skus']) - 10} more"
            ctk.CTkLabel(existing_frame, text=sku_text, font=ctk.CTkFont(size=11), 
                        text_color="#92400E").pack(anchor="w", padx=15, pady=(0, 10))
        
        # Sample Products Section
        if preview_data['sample_products']:
            sample_frame = ctk.CTkFrame(main_frame)
            sample_frame.pack(fill="both", expand=True, pady=(0, 15))
            ctk.CTkLabel(sample_frame, text="📦 Sample Products (First 10)", 
                        font=ctk.CTkFont(size=13, weight="bold")).pack(anchor="w", padx=15, pady=(10, 5))
            
            # Sample treeview
            sample_tree = ttk.Treeview(sample_frame, columns=("Name", "Category", "Brand", "Stock", "Cost", "Price"), 
                                       show="headings", height=8)
            sample_tree.pack(side="left", fill="both", expand=True, padx=(15, 5), pady=(0, 10))
            
            scrollbar = ttk.Scrollbar(sample_frame, orient="vertical", command=sample_tree.yview)
            scrollbar.pack(side="right", fill="y", padx=(0, 15), pady=(0, 10))
            sample_tree.configure(yscrollcommand=scrollbar.set)
            
            # Configure columns
            for col in ("Name", "Category", "Brand", "Stock", "Cost", "Price"):
                sample_tree.heading(col, text=col)
                sample_tree.column(col, width=100 if col != "Name" else 200)
            
            # Insert sample data
            for prod in preview_data['sample_products']:
                sample_tree.insert("", "end", values=(
                    prod['name'][:40],
                    prod['category'],
                    prod['brand'],
                    prod['stock'],
                    f"₹{prod['cost_price']}",
                    f"₹{prod['price_normal']}"
                ))
        
        # Errors Section
        if preview_data['errors']:
            error_frame = ctk.CTkFrame(main_frame, fg_color="#FEE2E2")
            error_frame.pack(fill="x", pady=(0, 15))
            ctk.CTkLabel(error_frame, text=f"⚠️ Errors ({len(preview_data['errors'])})", 
                        font=ctk.CTkFont(size=13, weight="bold"), text_color="#DC2626").pack(anchor="w", padx=15, pady=(10, 5))
            for error in preview_data['errors'][:5]:
                ctk.CTkLabel(error_frame, text=f"  • {error}", font=ctk.CTkFont(size=10), 
                            text_color="#DC2626").pack(anchor="w", padx=15)
            if len(preview_data['errors']) > 5:
                ctk.CTkLabel(error_frame, text=f"  ... and {len(preview_data['errors']) - 5} more errors", 
                            font=ctk.CTkFont(size=10), text_color="#DC2626").pack(anchor="w", padx=15, pady=(0, 10))
        
        # Buttons
        button_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        button_frame.pack(fill="x", pady=(10, 0))
        
        result = {"confirmed": False}
        
        def on_confirm():
            result["confirmed"] = True
            preview_dialog.destroy()
        
        def on_cancel():
            result["confirmed"] = False
            preview_dialog.destroy()
        
        ctk.CTkButton(button_frame, text="✓ Proceed with Import", command=on_confirm,
                     fg_color="#10B981", hover_color="#059669",
                     width=180, height=45, font=ctk.CTkFont(size=13, weight="bold")).pack(side="left", padx=(0, 10))
        
        ctk.CTkButton(button_frame, text="✗ Cancel Import", command=on_cancel,
                     fg_color="#EF4444", hover_color="#DC2626",
                     width=150, height=45, font=ctk.CTkFont(size=13)).pack(side="left")
        
        # Wait for dialog
        self.dialog.wait_window(preview_dialog)
        return result["confirmed"]
    
    def _do_import(self, mappings, preview_data):
        """Actually perform the import"""
        imported_count = 0
        errors = []
        
        for row_idx, row_data in enumerate(self.excel_data):
            try:
                # Extract values
                values = {}
                for field_key, col_idx in mappings.items():
                    val = row_data[col_idx] if col_idx < len(row_data) else ""
                    field_info = self.db_fields[field_key]
                    if field_info['type'] == 'number':
                        try:
                            val = float(val) if val else (field_info.get('default', 0))
                        except:
                            val = field_info.get('default', 0)
                    values[field_key] = val
                
                # Skip if missing required fields
                skip_row = False
                for field_key, field_info in self.db_fields.items():
                    if field_info['required'] and not values.get(field_key):
                        skip_row = True
                        break
                
                if skip_row:
                    continue
                
                # Handle category/sub_category/brand/sub_brand
                category_id = self.get_or_create_category(values.get('category', ''))
                sub_category_id = self.get_or_create_sub_category(values.get('sub_category', ''), category_id)
                brand_id = self.get_or_create_brand(values.get('brand', ''))
                sub_brand_id = self.get_or_create_sub_brand(values.get('sub_brand', ''), brand_id)
                
                # Generate SKU if not provided
                sku = values.get('sku', '').strip()
                if not sku:
                    sku = SKUGenerator.generate_sku(category_id, brand_id)
                
                # Check if SKU already exists
                existing = db.execute_query("SELECT product_id FROM products WHERE sku = ?", (sku,))
                if existing:
                    db.execute_update(
                        """UPDATE products SET name=?, category_id=?, sub_category_id=?, 
                           brand_id=?, sub_brand_id=?, description=?, stock=?, 
                           cost_price=?, price_normal=?, price_workshop=?, reorder_level=?
                           WHERE sku=?""",
                        (values.get('name', ''), category_id, sub_category_id, brand_id, sub_brand_id,
                         values.get('description', ''), values.get('stock', 0),
                         values.get('cost_price', 0), values.get('price_normal', 0),
                         values.get('price_workshop', 0), values.get('reorder_level', 10), sku)
                    )
                else:
                    db.execute_insert(
                        """INSERT INTO products (name, sku, category_id, sub_category_id, 
                           brand_id, sub_brand_id, description, stock, cost_price, 
                           price_normal, price_workshop, reorder_level, is_active)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)""",
                        (values.get('name', ''), sku, category_id, sub_category_id, brand_id, sub_brand_id,
                         values.get('description', ''), values.get('stock', 0),
                         values.get('cost_price', 0), values.get('price_normal', 0),
                         values.get('price_workshop', 0), values.get('reorder_level', 10))
                    )
                
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_idx + 2}: {str(e)}")
        
        # Show results
        if errors:
            error_msg = "Import completed with some errors:\n\n" + "\n".join(errors[:10])
            if len(errors) > 10:
                error_msg += f"\n... and {len(errors) - 10} more errors"
            messagebox.showwarning("Import Results", error_msg)
        
        return imported_count
    
    def get_or_create_category(self, category_name):
        """Get or create category by name"""
        if not category_name:
            return None
        
        category = db.execute_query("SELECT category_id FROM categories WHERE LOWER(category_name) = LOWER(?)", 
                                    (category_name,))
        if category:
            return category[0]['category_id']
        
        return db.execute_insert("INSERT INTO categories (category_name) VALUES (?)", (category_name,))
    
    def get_or_create_sub_category(self, sub_category_name, category_id):
        """Get or create sub category by name"""
        if not sub_category_name or not category_id:
            return None
        
        sub_category = db.execute_query(
            "SELECT sub_category_id FROM sub_categories WHERE LOWER(sub_category_name) = LOWER(?) AND category_id = ?",
            (sub_category_name, category_id))
        if sub_category:
            return sub_category[0]['sub_category_id']
        
        return db.execute_insert(
            "INSERT INTO sub_categories (sub_category_name, category_id) VALUES (?, ?)",
            (sub_category_name, category_id))
    
    def get_or_create_brand(self, brand_name):
        """Get or create brand by name"""
        if not brand_name:
            return None
        
        brand = db.execute_query("SELECT brand_id FROM brands WHERE LOWER(brand_name) = LOWER(?)", 
                                 (brand_name,))
        if brand:
            return brand[0]['brand_id']
        
        return db.execute_insert("INSERT INTO brands (brand_name) VALUES (?)", (brand_name,))
    
    def get_or_create_sub_brand(self, sub_brand_name, brand_id):
        """Get or create sub brand by name"""
        if not sub_brand_name or not brand_id:
            return None
        
        sub_brand = db.execute_query(
            "SELECT sub_brand_id FROM sub_brands WHERE LOWER(sub_brand_name) = LOWER(?) AND brand_id = ?",
            (sub_brand_name, brand_id))
        if sub_brand:
            return sub_brand[0]['sub_brand_id']
        
        return db.execute_insert(
            "INSERT INTO sub_brands (sub_brand_name, brand_id) VALUES (?, ?)",
            (sub_brand_name, brand_id))
    
    def on_mapping_changed(self, field_key, selected_value):
        """Called when a mapping dropdown is changed - refreshes the preview"""
        # Update the preview to show the new mapping
        self.update_preview()
    
    def cancel(self):
        """Cancel import"""
        self.result = None
        self.dialog.destroy()

    def show_sheet_selector(self, wb, sheet_names):
        """Show sheet selector dialog with navigation to preview and back"""
        self._selected_sheet_result = None
        
        # Calculate responsive dimensions
        parent_width = self.dialog.winfo_width()
        parent_height = self.dialog.winfo_height()
        if parent_width < 100:
            parent_width = 1000
        if parent_height < 100:
            parent_height = 700
        
        dialog_width = int(min(parent_width * 0.7, 700))
        dialog_height = int(min(parent_height * 0.85, 750))
        
        # Create main sheet selector dialog
        selector_dialog = tk.Toplevel(self.dialog)
        selector_dialog.title("Select Sheet")
        selector_dialog.geometry(f"{dialog_width}x{dialog_height}")
        selector_dialog.transient(self.dialog)
        selector_dialog.grab_set()
        selector_dialog.configure(bg="#2D2D2D")
        
        # Center on parent
        selector_dialog.update_idletasks()
        x = self.dialog.winfo_x() + (self.dialog.winfo_width() // 2) - (dialog_width // 2)
        y = self.dialog.winfo_y() + (self.dialog.winfo_height() // 2) - (dialog_height // 2)
        x = max(0, min(x, selector_dialog.winfo_screenwidth() - dialog_width))
        y = max(0, min(y, selector_dialog.winfo_screenheight() - dialog_height))
        selector_dialog.geometry(f"{dialog_width}x{dialog_height}+{x}+{y}")
        
        # Header
        header_frame = ctk.CTkFrame(selector_dialog, corner_radius=0, fg_color="#3B82F6")
        header_frame.pack(fill="x", pady=0)
        ctk.CTkLabel(header_frame, text="📑 Select Sheet to Import", 
                    font=ctk.CTkFont(size=16, weight="bold"), text_color="white").pack(pady=15)
        
        ctk.CTkLabel(selector_dialog, text=f"This Excel file has {len(sheet_names)} sheets. Select a sheet:", 
                    font=ctk.CTkFont(size=12), text_color="gray").pack(pady=(10, 5))
        
        # Sheet list frame
        list_frame = ctk.CTkFrame(selector_dialog, fg_color="#2D2D2D")
        list_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Create listbox for ALL sheets
        sheet_listbox = tk.Listbox(list_frame, font=("Segoe UI", 13), 
                                   bg="#2D2D2D", fg="white",
                                   selectbackground="#3B82F6", 
                                   selectforeground="white",
                                   borderwidth=1, highlightthickness=1,
                                   height=25, width=50)
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=sheet_listbox.yview)
        sheet_listbox.configure(yscrollcommand=scrollbar.set)
        
        sheet_listbox.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Populate ALL sheet names
        for name in sheet_names:
            sheet_listbox.insert(tk.END, name)
        
        def on_select_clicked():
            """Handle Select button click"""
            selection = sheet_listbox.curselection()
            if not selection:
                messagebox.showwarning("Warning", "Please select a sheet first")
                return
            
            sheet_name = sheet_listbox.get(selection[0])
            
            # Hide selector and show preview
            selector_dialog.withdraw()
            preview_result = self.show_sheet_preview(wb, sheet_name, selector_dialog)
            
            if preview_result == "SELECT":
                self._selected_sheet_result = sheet_name
                selector_dialog.destroy()
            elif preview_result == "BACK":
                # Show selector again
                selector_dialog.deiconify()
        
        # Button frame
        btn_frame = ctk.CTkFrame(selector_dialog, fg_color="transparent")
        btn_frame.pack(fill="x", pady=15, padx=20)
        
        ctk.CTkButton(btn_frame, text="Preview Sheet", command=on_select_clicked,
                     fg_color="#3B82F6", hover_color="#2563EB",
                     width=150, height=40, font=ctk.CTkFont(size=12, weight="bold")).pack(side="left")
        
        ctk.CTkButton(btn_frame, text="Cancel", command=lambda: selector_dialog.destroy(),
                     fg_color="#6B7280", hover_color="#4B5563",
                     width=100, height=40).pack(side="right")
        
        self.dialog.wait_window(selector_dialog)
        return self._selected_sheet_result

    def show_sheet_preview(self, wb, sheet_name, parent_dialog):
        """Show detailed preview of selected sheet with BACK button - responsive and shows more data"""
        result = None
        
        # Calculate responsive dimensions
        parent_width = self.dialog.winfo_width()
        parent_height = self.dialog.winfo_height()
        if parent_width < 100:
            parent_width = 1000
        if parent_height < 100:
            parent_height = 700
        
        dialog_width = int(min(parent_width * 0.9, 1100))
        dialog_height = int(min(parent_height * 0.9, 750))
        
        preview_dialog = tk.Toplevel(self.dialog)
        preview_dialog.title(f"Preview: {sheet_name}")
        preview_dialog.geometry(f"{dialog_width}x{dialog_height}")
        preview_dialog.transient(self.dialog)
        preview_dialog.grab_set()
        preview_dialog.configure(bg="#2D2D2D")
        
        # Center on parent
        preview_dialog.update_idletasks()
        x = self.dialog.winfo_x() + (self.dialog.winfo_width() // 2) - (dialog_width // 2)
        y = self.dialog.winfo_y() + (self.dialog.winfo_height() // 2) - (dialog_height // 2)
        x = max(0, min(x, preview_dialog.winfo_screenwidth() - dialog_width))
        y = max(0, min(y, preview_dialog.winfo_screenheight() - dialog_height))
        preview_dialog.geometry(f"{dialog_width}x{dialog_height}+{x}+{y}")
        
        # Header
        header_frame = ctk.CTkFrame(preview_dialog, corner_radius=0, fg_color="#10B981")
        header_frame.pack(fill="x", pady=0)
        ctk.CTkLabel(header_frame, text=f"📄 {sheet_name}", 
                    font=ctk.CTkFont(size=16, weight="bold"), text_color="white").pack(pady=15)
        
        # Content frame
        content_frame = ctk.CTkFrame(preview_dialog, fg_color="#2D2D2D")
        content_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Show more columns and rows in preview
        ctk.CTkLabel(content_frame, text="Sheet Content Preview (First 20 rows, first 20 columns):", 
                    font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", pady=(0, 5))
        
        # Preview text widget with more content
        preview_text = tk.Text(content_frame, font=("Courier", 10), 
                              bg="#1E1E1E", fg="white",
                              borderwidth=1, highlightthickness=1,
                              height=28, width=100, wrap="none")
        scrollbar_y = ttk.Scrollbar(content_frame, orient="vertical", command=preview_text.yview)
        scrollbar_x = ttk.Scrollbar(content_frame, orient="horizontal", command=preview_text.xview)
        preview_text.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        
        preview_text.pack(side="top", fill="both", expand=True)
        scrollbar_y.pack(side="right", fill="y")
        scrollbar_x.pack(side="bottom", fill="x")
        
        # Load and display sheet content - show more data
        try:
            sheet = wb[sheet_name]
            preview_lines = []
            
            # Get first 20 rows, 20 columns for more comprehensive preview
            for row_num in range(1, min(21, sheet.max_row + 1)):
                row = sheet[row_num]
                row_data = []
                for cell in row[:20]:  # First 20 columns
                    val = cell.value
                    if val is not None:
                        # Truncate long values
                        text = str(val)[:30]
                        row_data.append(text)
                    else:
                        row_data.append("")
                preview_lines.append(f"Row {row_num:2}: {' | '.join(row_data)}")
            
            preview_text.insert("1.0", "\n".join(preview_lines))
            preview_text.configure(state="disabled")
        except Exception as e:
            preview_text.insert("1.0", f"Error loading preview: {str(e)}")
            preview_text.configure(state="disabled")
        
        # Button frame
        btn_frame = ctk.CTkFrame(preview_dialog, fg_color="transparent")
        btn_frame.pack(fill="x", pady=15, padx=20)
        
        def on_back():
            nonlocal result
            result = "BACK"
            preview_dialog.destroy()
        
        def on_select():
            nonlocal result
            result = "SELECT"
            preview_dialog.destroy()
        
        ctk.CTkButton(btn_frame, text="← Back", command=on_back,
                     fg_color="#6B7280", hover_color="#4B5563",
                     width=120, height=40, font=ctk.CTkFont(size=12)).pack(side="left", padx=(0, 10))
        
        ctk.CTkButton(btn_frame, text="Select This Sheet", command=on_select,
                     fg_color="#10B981", hover_color="#059669", 
                     width=180, height=40, font=ctk.CTkFont(size=12, weight="bold")).pack(side="right")
        
        self.dialog.wait_window(preview_dialog)
        return result


class ProductDialog:
    def __init__(self, parent, title, product_data=None):
        self.parent = parent
        self.result = None
        self._barcode_server = None
        self._server_url = None
        
        # Create dialog
        self.dialog = ctk.CTkToplevel(parent)
        self.dialog.title(title)
        self.dialog.geometry("500x750")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # Center dialog on parent window
        self.dialog.update_idletasks()
        parent.update_idletasks()
        
        parent_x = parent.winfo_rootx()
        parent_y = parent.winfo_rooty()
        parent_width = parent.winfo_width()
        parent_height = parent.winfo_height()
        
        dialog_width = 500
        dialog_height = 750
        
        x = parent_x + (parent_width // 2) - (dialog_width // 2)
        y = parent_y + (parent_height // 2) - (dialog_height // 2)
        
        self.dialog.geometry(f"{dialog_width}x{dialog_height}+{x}+{y}")
        
        # Load categories and brands
        self.load_categories_and_brands()
        
        # Create form
        self.create_form(product_data)
        
        # Wait for dialog to close
        self.dialog.wait_window()
    
    def load_categories_and_brands(self):
        """Load categories and brands for dropdowns"""
        self.categories = db.execute_query("SELECT category_id, category_name FROM categories ORDER BY category_name")
        self.brands = db.execute_query("SELECT brand_id, brand_name FROM brands ORDER BY brand_name")
        self.sub_categories = db.execute_query("SELECT sub_category_id, sub_category_name, category_id FROM sub_categories ORDER BY sub_category_name")
        self.sub_brands = db.execute_query("SELECT sub_brand_id, sub_brand_name, brand_id FROM sub_brands ORDER BY sub_brand_name")
    
    def create_form(self, product_data):
        """Create product form"""
        # Main frame
        main_frame = ctk.CTkScrollableFrame(self.dialog)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Form fields
        # Name
        ctk.CTkLabel(main_frame, text="Product Name:", font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", pady=(0, 5))
        self.name_entry = ctk.CTkEntry(main_frame, width=400)
        self.name_entry.pack(anchor="w", pady=(0, 10))
        
        # Category with searchable dropdown
        ctk.CTkLabel(main_frame, text="Category:", font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", pady=(0, 5))
        self.category_dropdown = SearchableDropdown(main_frame, self.categories, 'category_name', 'category_id', 
                                                     placeholder="Select Category...", width=400, height=35,
                                                     command=self.on_category_selected,
                                                     add_command=self.add_category_from_dropdown)
        self.category_dropdown.pack(fill="x", pady=(0, 10))
        
        # Sub Category with searchable dropdown - initialize with all sub categories
        ctk.CTkLabel(main_frame, text="Sub Category:", font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", pady=(0, 5))
        self.sub_category_dropdown = SearchableDropdown(main_frame, self.sub_categories, 'sub_category_name', 'sub_category_id',
                                                         placeholder="Select Sub Category...", width=400, height=35,
                                                         add_command=self.add_sub_category_from_dropdown)
        self.sub_category_dropdown.pack(fill="x", pady=(0, 10))
        
        # Brand with searchable dropdown
        ctk.CTkLabel(main_frame, text="Brand:", font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", pady=(0, 5))
        self.brand_dropdown = SearchableDropdown(main_frame, self.brands, 'brand_name', 'brand_id',
                                                  placeholder="Select Brand...", width=400, height=35,
                                                  command=self.on_brand_selected,
                                                  add_command=self.add_brand_from_dropdown)
        self.brand_dropdown.pack(fill="x", pady=(0, 10))
        
        # Sub Brand with searchable dropdown - initialize with all sub brands
        ctk.CTkLabel(main_frame, text="Sub Brand:", font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", pady=(0, 5))
        self.sub_brand_dropdown = SearchableDropdown(main_frame, self.sub_brands, 'sub_brand_name', 'sub_brand_id',
                                                      placeholder="Select Sub Brand...", width=400, height=35,
                                                      add_command=self.add_sub_brand_from_dropdown)
        self.sub_brand_dropdown.pack(fill="x", pady=(0, 10))
        
        # Description
        ctk.CTkLabel(main_frame, text="Description:", font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", pady=(0, 5))
        self.description_entry = ctk.CTkEntry(main_frame, width=400)
        self.description_entry.pack(anchor="w", pady=(0, 10))
        
        # Barcode
        ctk.CTkLabel(main_frame, text="Barcode:", font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", pady=(0, 5))
        barcode_row = ctk.CTkFrame(main_frame, fg_color="transparent")
        barcode_row.pack(fill="x", pady=(0, 5))
        self.barcode_entry = ctk.CTkEntry(barcode_row, width=220, placeholder_text="Scan or type barcode...")
        self.barcode_entry.pack(side="left", padx=(0, 5))
        self.barcode_entry.bind("<Return>", self._on_barcode_scanned)
        self.barcode_entry.bind("<KeyRelease>", self._on_barcode_edited)
        self.cam_btn = ctk.CTkButton(
            barcode_row, text="Camera", width=65,
            command=self._start_camera_scanner
        )
        self.cam_btn.pack(side="left", padx=(0, 3))
        self.phone_btn = ctk.CTkButton(
            barcode_row, text="Phone", width=65,
            command=self._start_phone_scanner
        )
        self.phone_btn.pack(side="left")
        # Second row: preview + action buttons
        barcode_row2 = ctk.CTkFrame(main_frame, fg_color="transparent")
        barcode_row2.pack(fill="x", pady=(0, 5))
        self.barcode_preview = ctk.CTkLabel(barcode_row2, text="")
        self.barcode_preview.pack(side="left", padx=(0, 10))
        btn_frame = ctk.CTkFrame(barcode_row2, fg_color="transparent")
        btn_frame.pack(side="left")
        self.gen_btn = ctk.CTkButton(
            btn_frame, text="Generate", width=80,
            command=self._generate_barcode
        )
        self.gen_btn.pack(side="left", padx=(0, 5))
        self.print_btn = ctk.CTkButton(
            btn_frame, text="Print", width=70,
            command=self._print_barcode
        )
        self.print_btn.pack(side="left")

        self.qr_print_btn = ctk.CTkButton(
            btn_frame, text="Print Barcode (Template)", width=165,
            command=self._print_qr_in_template
        )
        self.qr_print_btn.pack(side="left", padx=(8, 0))
        self.scan_info = ctk.CTkLabel(main_frame, text="", font=ctk.CTkFont(size=11))
        self.scan_info.pack(anchor="w", pady=(0, 10))

        self.dialog.protocol("WM_DELETE_WINDOW", self._on_dialog_close)
        
        # Stock
        ctk.CTkLabel(main_frame, text="Stock Quantity:", font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", pady=(0, 5))
        self.stock_entry = ctk.CTkEntry(main_frame, width=400)
        self.stock_entry.pack(anchor="w", pady=(0, 10))
        
        # Cost Price
        ctk.CTkLabel(main_frame, text="Cost Price:", font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", pady=(0, 5))
        self.cost_entry = ctk.CTkEntry(main_frame, width=400)
        self.cost_entry.pack(anchor="w", pady=(0, 10))
        
        # Normal Price
        ctk.CTkLabel(main_frame, text="Normal Price:", font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", pady=(0, 5))
        self.normal_price_entry = ctk.CTkEntry(main_frame, width=400)
        self.normal_price_entry.pack(anchor="w", pady=(0, 10))
        
        # Workshop Price
        ctk.CTkLabel(main_frame, text="Workshop Price:", font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", pady=(0, 5))
        self.workshop_price_entry = ctk.CTkEntry(main_frame, width=400)
        self.workshop_price_entry.pack(anchor="w", pady=(0, 10))
        
        # Reorder Level
        ctk.CTkLabel(main_frame, text="Reorder Level:", font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", pady=(0, 5))
        self.reorder_entry = ctk.CTkEntry(main_frame, width=400)
        self.reorder_entry.pack(anchor="w", pady=(0, 10))
        
        # Fill form if editing
        if product_data:
            # Handle SQLite Row objects by accessing columns directly
            self.name_entry.insert(0, product_data['name'] if product_data['name'] else '')
            self.description_entry.insert(0, product_data['description'] if product_data['description'] else '')
            self.barcode_entry.insert(0, product_data.get('barcode', '') or '')
            self.stock_entry.insert(0, str(product_data['stock']) if product_data['stock'] else '')
            self.cost_entry.insert(0, str(product_data['cost_price']) if product_data['cost_price'] else '')
            self.normal_price_entry.insert(0, str(product_data['price_normal']) if product_data['price_normal'] else '')
            self.workshop_price_entry.insert(0, str(product_data['price_workshop']) if product_data['price_workshop'] else '')
            self.reorder_entry.insert(0, str(product_data['reorder_level']) if product_data['reorder_level'] else '')
            
            # Set dropdowns
            if product_data['category_id']:
                for cat in self.categories:
                    if cat['category_id'] == product_data['category_id']:
                        self.category_dropdown.set(cat['category_name'], cat['category_id'])
                        # Update sub categories based on selected category
                        self._update_sub_categories(cat['category_id'])
                        break
            
            if product_data['sub_category_id']:
                for sc in self.sub_categories:
                    if sc['sub_category_id'] == product_data['sub_category_id']:
                        self.sub_category_dropdown.set(sc['sub_category_name'], sc['sub_category_id'])
                        break
            
            if product_data['brand_id']:
                for brand in self.brands:
                    if brand['brand_id'] == product_data['brand_id']:
                        self.brand_dropdown.set(brand['brand_name'], brand['brand_id'])
                        # Update sub brands based on selected brand
                        self._update_sub_brands(brand['brand_id'])
                        break
            
            if product_data['sub_brand_id']:
                for sb in self.sub_brands:
                    if sb['sub_brand_id'] == product_data['sub_brand_id']:
                        self.sub_brand_dropdown.set(sb['sub_brand_name'], sb['sub_brand_id'])
                        break
        
        # Buttons
        button_frame = ctk.CTkFrame(main_frame)
        button_frame.pack(fill="x", pady=(20, 0))
        
        save_btn = ctk.CTkButton(
            button_frame,
            text="Save",
            command=self.save_product,
            fg_color=("#22C55E", "#16A34A"),
            hover_color=("#16A34A", "#15803D"),
            width=100
        )
        save_btn.pack(side="left", padx=(0, 10))
        
        cancel_btn = ctk.CTkButton(
            button_frame,
            text="Cancel",
            command=self.dialog.destroy,
            width=100
        )
        cancel_btn.pack(side="left")
    
    def on_category_selected(self):
        """Handle category selection - update sub categories"""
        category_id = self.category_dropdown.get()
        if category_id:
            self._update_sub_categories(category_id)
        else:
            self.sub_category_dropdown.set_items([])
            self.sub_category_dropdown.clear()
    
    def on_brand_selected(self):
        """Handle brand selection - update sub brands"""
        brand_id = self.brand_dropdown.get()
        if brand_id:
            self._update_sub_brands(brand_id)
        else:
            self.sub_brand_dropdown.set_items([])
            self.sub_brand_dropdown.clear()
    
    def _update_sub_categories(self, category_id):
        """Update sub categories based on selected category"""
        filtered_sub_categories = [sc for sc in self.sub_categories if sc['category_id'] == category_id]
        self.sub_category_dropdown.set_items(filtered_sub_categories)
        self.sub_category_dropdown.clear()
    
    def _update_sub_brands(self, brand_id):
        """Update sub brands based on selected brand"""
        filtered_sub_brands = [sb for sb in self.sub_brands if sb['brand_id'] == brand_id]
        self.sub_brand_dropdown.set_items(filtered_sub_brands)
        self.sub_brand_dropdown.clear()

    def add_category_from_dropdown(self):
        """Add new category from dropdown"""
        dialog = CategoryDialog(self.dialog, "Add Category")
        if dialog.result:
            try:
                # Insert category
                category_id = db.execute_insert(
                    "INSERT INTO categories (category_name, description) VALUES (?, ?)",
                    (dialog.result['name'], dialog.result['description'])
                )
                # Reload categories
                self.categories = db.execute_query("SELECT category_id, category_name FROM categories ORDER BY category_name")
                # Update dropdown
                self.category_dropdown.set_items(self.categories)
                # Select the new category
                self.category_dropdown.set(dialog.result['name'], category_id)
                # Trigger category selection to update sub categories
                self.on_category_selected()
                messagebox.showinfo("Success", "Category added successfully!")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to add category: {e}")

    def add_brand_from_dropdown(self):
        """Add new brand from dropdown"""
        dialog = BrandDialog(self.dialog, "Add Brand")
        if dialog.result:
            try:
                # Insert brand
                brand_id = db.execute_insert(
                    "INSERT INTO brands (brand_name, description) VALUES (?, ?)",
                    (dialog.result['name'], dialog.result['description'])
                )
                # Reload brands
                self.brands = db.execute_query("SELECT brand_id, brand_name FROM brands ORDER BY brand_name")
                # Update dropdown
                self.brand_dropdown.set_items(self.brands)
                # Select the new brand
                self.brand_dropdown.set(dialog.result['name'], brand_id)
                # Trigger brand selection to update sub brands
                self.on_brand_selected()
                messagebox.showinfo("Success", "Brand added successfully!")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to add brand: {e}")

    def add_sub_category_from_dropdown(self):
        """Add new sub category from dropdown"""
        # Get currently selected category
        category_id = self.category_dropdown.get()
        if not category_id:
            messagebox.showwarning("Warning", "Please select a category first.")
            return
        
        # Get category name for pre-filling
        category_name = self.category_dropdown.get_name()
        
        dialog = SubCategoryDialog(self.dialog, "Add Sub Category", 
                                   sub_category_data={'category_id': category_id, 'category_name': category_name})
        if dialog.result:
            try:
                # Insert sub category
                sub_category_id = db.execute_insert(
                    "INSERT INTO sub_categories (sub_category_name, category_id, description) VALUES (?, ?, ?)",
                    (dialog.result['name'], dialog.result['category_id'], dialog.result['description'])
                )
                # Reload sub categories
                self.sub_categories = db.execute_query("SELECT sub_category_id, sub_category_name, category_id FROM sub_categories ORDER BY sub_category_name")
                # Update dropdown with filtered items (includes the new one)
                filtered_sub_categories = [sc for sc in self.sub_categories if sc['category_id'] == category_id]
                self.sub_category_dropdown.set_items(filtered_sub_categories)
                # Select the new sub category immediately
                self.sub_category_dropdown.set(dialog.result['name'], sub_category_id)
                messagebox.showinfo("Success", "Sub Category added successfully!")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to add sub category: {e}")

    def add_sub_brand_from_dropdown(self):
        """Add new sub brand from dropdown"""
        # Get currently selected brand
        brand_id = self.brand_dropdown.get()
        if not brand_id:
            messagebox.showwarning("Warning", "Please select a brand first.")
            return
        
        # Get brand name for pre-filling
        brand_name = self.brand_dropdown.get_name()
        
        dialog = SubBrandDialog(self.dialog, "Add Sub Brand",
                               sub_brand_data={'brand_id': brand_id, 'brand_name': brand_name})
        if dialog.result:
            try:
                # Insert sub brand
                sub_brand_id = db.execute_insert(
                    "INSERT INTO sub_brands (sub_brand_name, brand_id, description) VALUES (?, ?, ?)",
                    (dialog.result['name'], dialog.result['brand_id'], dialog.result['description'])
                )
                # Reload sub brands
                self.sub_brands = db.execute_query("SELECT sub_brand_id, sub_brand_name, brand_id FROM sub_brands ORDER BY sub_brand_name")
                # Update dropdown with filtered items (includes the new one)
                filtered_sub_brands = [sb for sb in self.sub_brands if sb['brand_id'] == brand_id]
                self.sub_brand_dropdown.set_items(filtered_sub_brands)
                # Select the new sub brand immediately
                self.sub_brand_dropdown.set(dialog.result['name'], sub_brand_id)
                messagebox.showinfo("Success", "Sub Brand added successfully!")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to add sub brand: {e}")

    def save_product(self):
        """Save product data"""
        try:
            # Validate required fields
            if not self.name_entry.get():
                messagebox.showerror("Error", "Product name is required")
                return
            
            # Get category and brand IDs from dropdowns
            category_id = self.category_dropdown.get()
            sub_category_id = self.sub_category_dropdown.get()
            brand_id = self.brand_dropdown.get()
            sub_brand_id = self.sub_brand_dropdown.get()
            
            # Validate numeric fields
            stock = int(self.stock_entry.get() or 0)
            cost_price = float(self.cost_entry.get() or 0)
            price_normal = float(self.normal_price_entry.get() or 0)
            price_workshop = float(self.workshop_price_entry.get() or 0)
            reorder_level = int(self.reorder_entry.get() or 10)
            
            # Data validation - prevent negative stock
            if stock < 0:
                messagebox.showerror("Validation Error", "Stock quantity cannot be negative!")
                return
            if cost_price < 0:
                messagebox.showerror("Validation Error", "Cost price cannot be negative!")
                return
            if price_normal < 0:
                messagebox.showerror("Validation Error", "Normal price cannot be negative!")
                return
            if price_workshop < 0:
                messagebox.showerror("Validation Error", "Workshop price cannot be negative!")
                return
            if reorder_level < 0:
                messagebox.showerror("Validation Error", "Reorder level cannot be negative!")
                return
            
            self.result = {
                'name': self.name_entry.get(),
                'category_id': category_id,
                'sub_category_id': sub_category_id,
                'brand_id': brand_id,
                'sub_brand_id': sub_brand_id,
                'description': self.description_entry.get(),
                'barcode': self.barcode_entry.get().strip() or None,
                'stock': stock,
                'cost_price': cost_price,
                'price_normal': price_normal,
                'price_workshop': price_workshop,
                'reorder_level': reorder_level
            }
            
            self._stop_barcode_server()
            self.dialog.destroy()
            
        except ValueError:
            messagebox.showerror("Error", "Please enter valid numeric values for prices and quantities")
        except Exception as e:
            messagebox.showerror("Error", f"Error saving product: {e}")


    def _on_barcode_scanned(self, event=None):
        barcode = self.barcode_entry.get().strip()
        if barcode:
            self.barcode_entry.configure(border_color="green")
            self.parent.after(2000, lambda: self._reset_barcode_border())

    def _reset_barcode_border(self):
        try:
            if self.dialog.winfo_exists():
                self.barcode_entry.configure(border_color="#565b5e")
        except Exception:
            pass

    def _start_camera_scanner(self):
        """
        Start barcode camera scanner.

        NOTE: CameraBarcodeReader import is currently commented out in this file.
        Implement a safe behavior to avoid crashing when the user clicks Camera.
        """
        try:
            # If CameraBarcodeReader exists/import is enabled, prefer it.
            from utils.camera_scanner import CameraBarcodeReader  # type: ignore
            reader = CameraBarcodeReader(self.dialog, self._on_camera_barcode)
            reader.start()
        except Exception:
            messagebox.showinfo(
                "Camera Scanner Unavailable",
                "Camera scanning is not available in this build.\n"
                "Use Phone scanning or type the barcode manually."
            )

    def _on_camera_barcode(self, code):
        try:
            if not self.dialog.winfo_exists():
                return
            self.barcode_entry.delete(0, "end")
            self.barcode_entry.insert(0, code)
            self.barcode_entry.configure(border_color="green")
            self.parent.after(2000, lambda: self._reset_barcode_border())
        except Exception:
            pass



    def _stop_barcode_server(self):
        if self._barcode_server:
            try:
                self._barcode_server.stop()
            except Exception:
                pass
            self._barcode_server = None
            self._server_url = None

    def _on_dialog_close(self):
        self._stop_barcode_server()
        self.dialog.destroy()

    def _start_phone_scanner(self):
        try:
            if self._barcode_server:
                self.scan_info.configure(
                    text=f"Open {self._server_url} on your phone (Chrome on Android recommended). Make sure both devices are on the same Wi-Fi.",
                    text_color="#60a5fa"
                )
                return
            self._barcode_server = BarcodeServer(port=8766)
            self._barcode_server.start(self._on_remote_scan)
            self._server_url = self._barcode_server.url
            self.scan_info.configure(
                text=f"Open {self._server_url} on your phone (Chrome on Android recommended). Same Wi-Fi required.",
                text_color="#60a5fa"
            )
        except Exception as e:
            self.scan_info.configure(text=f"Phone scanner error: {e}", text_color="#ef4444")

    def _on_remote_scan(self, code):
        self.parent.after(0, lambda: self._set_phone_barcode(code))

    def _set_phone_barcode(self, code):
        try:
            if not self.dialog.winfo_exists():
                return
            self.barcode_entry.delete(0, "end")
            self.barcode_entry.insert(0, code)
            self.barcode_entry.configure(border_color="green")
            self.parent.after(2000, lambda: self._reset_barcode_border())
        except Exception:
            pass

    def _on_barcode_edited(self, event=None):
        code = self.barcode_entry.get().strip()
        if code:
            self._show_barcode_preview(code)
        else:
            self.barcode_preview.configure(text="")

    def _generate_barcode(self):
        current = self.barcode_entry.get().strip()
        if current:
            if not messagebox.askyesno(
                "Barcode Already Generated",
                f"Barcode '{current}' already exists for this product.\n\nDo you want to generate a new one?"
            ):
                return
        existing = set()
        for row in db.execute_query("SELECT barcode FROM products WHERE barcode IS NOT NULL"):
            existing.add(row["barcode"])
        while True:
            code = "PRD" + "".join(random.choices(string.digits, k=8))
            if code not in existing:
                break
        self.barcode_entry.delete(0, "end")
        self.barcode_entry.insert(0, code)
        self._show_barcode_preview(code)
        self.scan_info.configure(text="Barcode generated successfully", text_color="#4ade80")

    def _show_barcode_preview(self, code):
        try:
            rv = io.BytesIO()
            Code128(code, writer=ImageWriter()).write(rv)
            rv.seek(0)
            img = Image.open(rv)
            img.thumbnail((160, 56))
            ctk_img = ctk.CTkImage(light_image=img, dark_image=img, size=img.size)
            self.barcode_preview.configure(image=ctk_img, text="")
        except Exception:
            pass

    def _print_barcode(self):
        code = self.barcode_entry.get().strip()
        name = self.name_entry.get().strip() or "Product"
        if not code:
            messagebox.showwarning("No Barcode", "Please generate or enter a barcode first.")
            return

        from reportlab.lib.pagesizes import inch
        from reportlab.pdfgen import canvas

        img_path = os.path.join(tempfile.gettempdir(), f"barcode_{code}.png")
        with open(img_path, "wb") as f:
            Code128(code, writer=ImageWriter()).write(f)

        pdf_path = os.path.join(tempfile.gettempdir(), f"barcode_{code}.pdf")
        c = canvas.Canvas(pdf_path, pagesize=(4 * inch, 2 * inch))
        c.setFont("Helvetica", 12)
        c.drawString(0.5 * inch, 1.5 * inch, f"Product: {name}")
        c.setFont("Helvetica", 10)
        c.drawString(0.5 * inch, 1.3 * inch, f"Barcode: {code}")
        c.drawImage(img_path, 0.5 * inch, 0.2 * inch, width=3 * inch, height=0.9 * inch)
        c.save()

        self.scan_info.configure(text=f"Barcode PDF saved: {pdf_path}", text_color="#60a5fa")
        os.startfile(pdf_path)

    def _show_qty_dialog(self):
        """Show a custom CTk dialog for print quantity - matches app styling"""
        dialog = ctk.CTkToplevel(self.parent)
        dialog.title("Print Quantity")
        dialog.geometry("400x200")
        dialog.resizable(False, False)
        dialog.attributes('-topmost', True)  # Keep on top
        dialog.transient(self.parent)
        
        # Center dialog on parent
        dialog.update_idletasks()
        self.parent.update_idletasks()
        
        parent_x = self.parent.winfo_rootx()
        parent_y = self.parent.winfo_rooty()
        parent_width = self.parent.winfo_width()
        parent_height = self.parent.winfo_height()
        
        x = parent_x + (parent_width // 2) - (200)
        y = parent_y + (parent_height // 2) - (100)
        dialog.geometry(f"400x200+{x}+{y}")
        
        # Make it modal
        dialog.grab_set()
        
        result = {"qty": None}
        
        # Title
        title_label = ctk.CTkLabel(
            dialog, 
            text="How many labels to print for this product?",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        title_label.pack(pady=(20, 10), padx=20)
        
        # Input frame
        input_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        input_frame.pack(pady=10, padx=20, fill="x")
        
        label = ctk.CTkLabel(input_frame, text="Quantity:", font=ctk.CTkFont(size=12))
        label.pack(side="left", padx=(0, 10))
        
        qty_entry = ctk.CTkEntry(
            input_frame, 
            font=ctk.CTkFont(size=12),
            width=150
        )
        qty_entry.pack(side="left", fill="x", expand=True)
        qty_entry.insert(0, "10")
        
        # Button frame
        button_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        button_frame.pack(pady=(15, 20), padx=20, fill="x")
        
        def on_ok():
            try:
                qty_val = int(qty_entry.get().strip())
                if qty_val < 1:
                    messagebox.showwarning("Invalid Input", "Please enter a number greater than 0")
                    return
                result["qty"] = qty_val
                dialog.destroy()
            except ValueError:
                messagebox.showerror("Invalid Input", "Please enter a valid number")
        
        def on_cancel():
            dialog.destroy()
        
        ok_btn = ctk.CTkButton(
            button_frame,
            text="OK",
            command=on_ok,
            font=ctk.CTkFont(size=12, weight="bold"),
            height=36
        )
        ok_btn.pack(side="left", padx=(0, 10), fill="x", expand=True)
        
        cancel_btn = ctk.CTkButton(
            button_frame,
            text="Cancel",
            command=on_cancel,
            font=ctk.CTkFont(size=12, weight="bold"),
            height=36,
            fg_color=("gray70", "gray30"),
            text_color=("gray10", "gray90")
        )
        cancel_btn.pack(side="left", fill="x", expand=True)
        
        # Focus on entry and select all
        qty_entry.focus()
        qty_entry.select_range(0, len(qty_entry.get()))
        
        # Wait for dialog to close
        dialog.wait_window()
        
        return result["qty"]

    def _print_qr_in_template(self):
        """Print Code128 barcodes (inside template boxes) + product name and price, for a user-selected quantity."""
        from utils.format_utils import get_currency_symbol
        
        raw_code = (self.barcode_entry.get() or "").strip()
        name = (self.name_entry.get() or "Product").strip() or "Product"
        price_str = (self.normal_price_entry.get() or "").strip()
        currency_symbol = get_currency_symbol()
        
        if not raw_code:
            messagebox.showwarning("No Barcode", "Please generate or enter a barcode first.")
            return

        # Quantity prompt: custom CTk dialog for consistent styling
        qty = self._show_qty_dialog()
        if not qty:
            return

        # Start box prompt + preview for user-selected "start printing" box
        start_box = self._show_start_box_dialog(qty)
        if not start_box:
            return

        # start_box is 1-based; convert to 0-based offset for placement loop
        offset = int(start_box) - 1

        # Product-name-only template (no price)

        payload = raw_code


        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas

        pdf_path = os.path.join(tempfile.gettempdir(), f"tpl_labels_{payload}.pdf")
        c = canvas.Canvas(pdf_path, pagesize=A4)
        w, h = A4

        # Template sizing matched to Eddy SI-84 A4 label sheet (46mm x 11mm x 84)
        # Each label: 46mm wide x 11mm tall
        box_w = 46 * mm  # 46mm width
        box_h = 11 * mm  # 11mm height
        margin_left = 5 * mm
        margin_bottom = 5 * mm
        gap_x = 2 * mm  # Small gap between columns
        gap_y = 2 * mm  # Small gap between rows

        # Layout grid (how many boxes fit) - no header, just margins
        footer_h = 15  # Minimal space for footer
        cols = max(1, int((w - 2*margin_left) // (box_w + gap_x)))
        available_height = h - margin_bottom - footer_h
        rows = max(1, int(available_height // (box_h + gap_y)))

        # Generate barcode image once and reuse for all boxes
        # NOTE: python-barcode PNGs include quiet-zone padding. To ensure the bars
        # never appear to "spill" outside the black template, we trim the image
        # down to the non-white content bounding box once here.
        barcode_png_path = os.path.join(tempfile.gettempdir(), f"tpl_barcode_{payload}.png")
        barcode_trimmed_png_path = os.path.join(tempfile.gettempdir(), f"tpl_barcode_{payload}_trim.png")
        rv = io.BytesIO()
        Code128(payload, writer=ImageWriter()).write(rv)
        rv.seek(0)
        with open(barcode_png_path, "wb") as f:
            f.write(rv.read())

        # Quiet-zone trim (white background -> trim to bbox of non-white pixels)
        try:
            img = Image.open(barcode_png_path).convert("RGB")
            # Treat near-white as background
            px = img.load()
            w_i, h_i = img.size
            left, top, right, bottom = w_i, h_i, -1, -1
            for yy in range(h_i):
                for xx in range(w_i):
                    r, g, b = px[xx, yy]
                    if (r, g, b) != (255, 255, 255):
                        left = min(left, xx)
                        top = min(top, yy)
                        right = max(right, xx)
                        bottom = max(bottom, yy)

            if right >= left and bottom >= top:
                # Crop to the bounding box, but remove the numeric text at the bottom
                # The numeric text is typically in the bottom 20% of the barcode image
                barcode_only_bottom = int(bottom - (bottom - top) * 0.20)
                cropped = img.crop((left, top, right + 1, barcode_only_bottom))
            else:
                cropped = img
            cropped.save(barcode_trimmed_png_path)
        except Exception:
            # If trimming fails, fall back to original image
            barcode_trimmed_png_path = barcode_png_path

        def draw_box(x0, y0):
            # Outer rectangle (white background with border)
            x0 = int(x0)
            y0 = int(y0)
            box_w_i = int(box_w)
            box_h_i = int(box_h)

            # Draw white background with black border
            c.setFillColorRGB(1, 1, 1)
            c.setLineWidth(0.5)
            c.setStrokeColorRGB(0, 0, 0)
            c.rect(x0, y0, box_w_i, box_h_i, stroke=1, fill=1)

            # For 46mm x 11mm labels: barcode at top, text at bottom
            pad_horiz = 1.5 * mm  # Padding left and right
            pad_vert = 0.05 * mm  # ~2px (at 96dpi) bottom padding

            gap_between = 0.7 * mm  # ~2px gap between barcode and text
            
            content_x0 = x0 + pad_horiz
            content_y0 = y0 + pad_vert
            content_w = max(1, box_w_i - 2 * pad_horiz)
            content_h = max(1, box_h_i - 2 * pad_vert)

            # Split: ~60% barcode at TOP, ~40% for product name and price at BOTTOM (minus gap)
            text_area_h = int(content_h * 0.40)
            barcode_area_h = int(content_h - text_area_h - gap_between)
            
            # TEXT AREA: positioned at BOTTOM of box
            text_start_y = content_y0
            
            if text_area_h > 1.5:
                c.setFillColorRGB(0, 0, 0)
                c.setFont("Helvetica-Bold", 7)  # Readable font size
                
                # Text baseline (vertical center of text area)
                text_y = int(text_start_y + text_area_h / 2)
                
                # Line 1: Product name (left side) - allow up to 20 characters
                name_display = name[:20] if len(name) <= 20 else name[:18] + ".."
                c.drawString(int(content_x0), text_y, name_display)
                
                # Line 2: Price with currency symbol (right side) - separate from product name
                if price_str:
                    try:
                        price_val = float(price_str)
                        price_display = f"{currency_symbol} {price_val:.0f}"
                        c.drawRightString(int(content_x0 + content_w), text_y, price_display)
                    except ValueError:
                        pass
            
            # BARCODE AREA: positioned at TOP of box with minimal margins for maximum size
            barcode_margin_h = 0.3 * mm  # Minimal horizontal margin
            barcode_margin_v = 0.2 * mm  # Minimal vertical margin
            
            barcode_w = int(content_w - 2 * barcode_margin_h)  # Use most of width
            barcode_h = int(barcode_area_h - 2 * barcode_margin_v)  # Use most of height
            
            # Position barcode at TOP with gap between barcode and text
            barcode_x = int(content_x0 + barcode_margin_h)
            barcode_y = int(content_y0 + text_area_h + gap_between + barcode_margin_v)

            c.drawImage(
                barcode_trimmed_png_path,
                barcode_x,
                barcode_y,
                width=barcode_w,
                height=barcode_h,
                preserveAspectRatio=False,  # Stretch to fill available space
                mask='auto'
            )

        from datetime import datetime
        total_boxes = qty
        printed = offset
        end_exclusive = offset + total_boxes

        for _ in range(total_boxes):
            if printed >= (cols * rows):
                # Footer + new page
                c.setFont("Helvetica", 8)
                c.setFillGray(0.4)
                c.drawString(int(margin_left), 10, datetime.now().strftime("%Y-%m-%d %H:%M"))
                c.showPage()
                printed = 0

            page_index = printed
            r = page_index // cols
            cidx = page_index % cols

            # Compute box position
            x0 = margin_left + cidx * (box_w + gap_x)
            y0_box = margin_bottom + (rows - 1 - r) * (box_h + gap_y)

            draw_box(x0, y0_box)
            printed += 1

            if printed >= end_exclusive:
                break


        # Footer on last page
        c.setFont("Helvetica", 8)
        c.setFillGray(0.4)
        c.drawString(int(margin_left), 10, datetime.now().strftime("%Y-%m-%d %H:%M"))

        c.showPage()
        c.save()

        self.scan_info.configure(text=f"Template Labels PDF saved: {pdf_path}", text_color="#60a5fa")
        os.startfile(pdf_path)

    def _show_start_box_dialog(self, qty):
        """Prompt user for which label box index to start printing from.

        Returns:
            int | None: 1-based start index, or None if cancelled.
        """
        if not qty or qty < 1:
            messagebox.showwarning("Invalid Quantity", "Quantity must be at least 1.")
            return None

        dialog = ctk.CTkToplevel(self.parent)
        dialog.title("Start Printing Box")
        dialog.geometry("420x260")
        dialog.resizable(False, False)
        dialog.attributes('-topmost', True)
        dialog.transient(self.parent)

        # Center dialog on parent
        dialog.update_idletasks()
        pw = self.parent.winfo_rootx()
        ph = self.parent.winfo_rooty()
        pwidth = self.parent.winfo_width()
        pheight = self.parent.winfo_height()
        x = pw + (pwidth // 2) - (210)
        y = ph + (pheight // 2) - (130)
        dialog.geometry(f"420x260+{x}+{y}")

        dialog.grab_set()

        result = {"start": None}

        header = ctk.CTkFrame(dialog, corner_radius=10, fg_color=("#3B82F6", "#2563EB"))
        header.pack(fill="x", padx=15, pady=(15, 10))
        ctk.CTkLabel(header, text="Choose Start Box", font=ctk.CTkFont(size=16, weight="bold"), text_color="white").pack(pady=12)

        ctk.CTkLabel(dialog, text=f"Labels to print: {qty}", font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", padx=20, pady=(10, 4))
        ctk.CTkLabel(dialog, text="Printing will begin from the selected box (1-based).", font=ctk.CTkFont(size=11), text_color=("gray30", "gray60")).pack(anchor="w", padx=20, pady=(0, 10))

        input_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        input_frame.pack(fill="x", padx=20, pady=(0, 12))

        ctk.CTkLabel(input_frame, text="Start at:", font=ctk.CTkFont(size=12, weight="bold")).pack(side="left")

        start_var = tk.StringVar(value="1")
        start_entry = ctk.CTkEntry(input_frame, textvariable=start_var, width=120)
        start_entry.pack(side="left", padx=(10, 0))
        start_entry.focus()
        start_entry.select_range(0, len(start_var.get()))

        # Button row
        btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_frame.pack(fill="x", padx=20, pady=(0, 12))

        def on_ok():
            try:
                v = int(start_var.get().strip())
                if v < 1 or v > qty:
                    messagebox.showwarning("Invalid Input", f"Enter a value between 1 and {qty}.")
                    return
                result["start"] = v
                dialog.destroy()
            except ValueError:
                messagebox.showwarning("Invalid Input", "Please enter a whole number.")

        def on_cancel():
            result["start"] = None
            dialog.destroy()

        ok_btn = ctk.CTkButton(btn_frame, text="✓ Start", command=on_ok, fg_color=("#10B981", "#059669"), hover_color=("#059669", "#047857"), height=36)
        ok_btn.pack(side="left", fill="x", expand=True, padx=(0, 10))

        cancel_btn = ctk.CTkButton(btn_frame, text="Cancel", command=on_cancel, fg_color=("#EF4444", "#B91C1C"), hover_color=("#DC2626", "#991B1B"), height=36)
        cancel_btn.pack(side="left", fill="x", expand=True)

        self.parent.wait_window(dialog)
        return result["start"]




class CategoryDialog:
    def __init__(self, parent, title, category_data=None):
        self.parent = parent
        self.result = None
        
        # Create dialog
        self.dialog = ctk.CTkToplevel(parent)
        self.dialog.title(title)
        self.dialog.geometry("500x400")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # Center dialog on parent window
        self.dialog.update_idletasks()
        parent.update_idletasks()
        
        # Calculate position to center on parent
        parent_x = parent.winfo_rootx()
        parent_y = parent.winfo_rooty()
        parent_width = parent.winfo_width()
        parent_height = parent.winfo_height()
        
        dialog_width = 500
        dialog_height = 400
        
        x = parent_x + (parent_width // 2) - (dialog_width // 2)
        y = parent_y + (parent_height // 2) - (dialog_height // 2)
        
        self.dialog.geometry(f"{dialog_width}x{dialog_height}+{x}+{y}")
        
        # Ensure dialog stays on top
        self.dialog.attributes("-topmost", True)
        self.dialog.focus_force()
        
        # Create form
        self.create_form(category_data)
        
        # Wait for dialog to close
        self.dialog.wait_window()
    
    def create_form(self, category_data):
        """Create category form with modern styling"""
        # Main container - scrollable to ensure all content is visible
        main_frame = ctk.CTkScrollableFrame(self.dialog)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Header with icon
        header_frame = ctk.CTkFrame(main_frame, corner_radius=10, fg_color=("#4CAF50", "#2E7D32"))
        header_frame.pack(fill="x", padx=10, pady=(10, 15))
        
        ctk.CTkLabel(header_frame, text="📁", font=ctk.CTkFont(size=24)).pack(pady=(10, 5))
        
        ctk.CTkLabel(
            header_frame,
            text="Category Information",
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color="white"
        ).pack(pady=(0, 10))
        
        # Form fields container
        fields_frame = ctk.CTkFrame(main_frame)
        fields_frame.pack(fill="x", padx=10, pady=(0, 15))
        
        # Name field
        name_label = ctk.CTkLabel(
            fields_frame,
            text="Category Name:",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        name_label.pack(anchor="w", padx=15, pady=(15, 5))
        
        self.name_entry = ctk.CTkEntry(
            fields_frame,
            placeholder_text="Enter category name...",
            font=ctk.CTkFont(size=14),
            height=35
        )
        self.name_entry.pack(fill="x", padx=15, pady=(0, 15))
        
        # Description field
        desc_label = ctk.CTkLabel(
            fields_frame,
            text="Description (Optional):",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        desc_label.pack(anchor="w", padx=15, pady=(0, 5))
        
        self.description_textbox = ctk.CTkTextbox(
            fields_frame,
            height=80,
            font=ctk.CTkFont(size=12)
        )
        self.description_textbox.pack(fill="x", padx=15, pady=(0, 15))
        
        # Fill form if editing
        if category_data:
            self.name_entry.insert(0, category_data.get('category_name', ''))
            if category_data.get('description'):
                self.description_textbox.insert("1.0", category_data['description'])
        
        # Buttons container
        button_frame = ctk.CTkFrame(main_frame)
        button_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        save_btn = ctk.CTkButton(
            button_frame,
            text="Save Category",
            command=self.save_category,
            width=140,
            height=40,
            fg_color=("#4CAF50", "#2E7D32"),
            hover_color=("#43A047", "#1B5E20"),
            font=ctk.CTkFont(size=13, weight="bold")
        )
        save_btn.pack(side="left", padx=15, pady=15)
        
        cancel_btn = ctk.CTkButton(
            button_frame,
            text="Cancel",
            command=self.cancel,
            width=100,
            height=40,
            fg_color=("#EF4444", "#B91C1C"),
            hover_color=("#DC2626", "#991B1B"),
            font=ctk.CTkFont(size=13, weight="bold")
        )
        cancel_btn.pack(side="right", padx=15, pady=15)
        
        # Focus on name entry
        self.dialog.after(100, lambda: self.name_entry.focus())
    
    def save_category(self):
        """Save category with validation"""
        try:
            name = self.name_entry.get().strip()
            if not name:
                messagebox.showerror("Error", "Category name is required.")
                return
            
            description = self.description_textbox.get("1.0", "end-1c").strip()
            
            self.result = {
                'name': name,
                'description': description or ""
            }
            
            self.dialog.destroy()
            
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")
    
    def cancel(self):
        """Cancel and close dialog"""
        self.result = None
        self.dialog.destroy()


class SubCategoryDialog:
    def __init__(self, parent, title, sub_category_data=None):
        self.parent = parent
        self.result = None
        
        # Create dialog
        self.dialog = ctk.CTkToplevel(parent)
        self.dialog.title(title)
        self.dialog.geometry("500x500")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # Center dialog on parent window
        self.dialog.update_idletasks()
        parent.update_idletasks()
        
        parent_x = parent.winfo_rootx()
        parent_y = parent.winfo_rooty()
        parent_width = parent.winfo_width()
        parent_height = parent.winfo_height()
        
        dialog_width = 500
        dialog_height = 500
        
        x = parent_x + (parent_width // 2) - (dialog_width // 2)
        y = parent_y + (parent_height // 2) - (dialog_height // 2)
        
        self.dialog.geometry(f"{dialog_width}x{dialog_height}+{x}+{y}")
        
        self.dialog.attributes("-topmost", True)
        self.dialog.focus_force()
        
        # Load categories
        self.categories = db.execute_query("SELECT category_id, category_name FROM categories ORDER BY category_name")
        
        # Create form
        self.create_form(sub_category_data)
        
        # Wait for dialog to close
        self.dialog.wait_window()
    
    def create_form(self, sub_category_data):
        """Create sub category form with modern styling"""
        main_frame = ctk.CTkScrollableFrame(self.dialog)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Header
        header_frame = ctk.CTkFrame(main_frame, corner_radius=10, fg_color=("#8BC34A", "#558B2F"))
        header_frame.pack(fill="x", padx=10, pady=(10, 15))
        
        ctk.CTkLabel(
            header_frame,
            text="Sub Category Information",
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color="white"
        ).pack(pady=15)
        
        # Form fields container
        fields_frame = ctk.CTkFrame(main_frame)
        fields_frame.pack(fill="x", padx=10, pady=(0, 15))
        
        # Category selection
        cat_label = ctk.CTkLabel(
            fields_frame,
            text="Parent Category:*",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        cat_label.pack(anchor="w", padx=15, pady=(15, 5))
        
        self.category_var = tk.StringVar()
        self.category_id_var = tk.IntVar()
        category_frame = ctk.CTkFrame(fields_frame, fg_color="transparent")
        category_frame.pack(fill="x", padx=15, pady=(0, 15))
        
        self.category_display = ctk.CTkEntry(category_frame, width=320, state="readonly", font=ctk.CTkFont(size=12))
        self.category_display.pack(side="left", padx=(0, 10))
        
        ctk.CTkButton(category_frame, text="Search", width=70, command=self.show_category_selector).pack(side="left")
        
        # Name field
        name_label = ctk.CTkLabel(
            fields_frame,
            text="Sub Category Name:*",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        name_label.pack(anchor="w", padx=15, pady=(0, 5))
        
        self.name_entry = ctk.CTkEntry(
            fields_frame,
            placeholder_text="Enter sub category name...",
            font=ctk.CTkFont(size=14),
            height=35
        )
        self.name_entry.pack(fill="x", padx=15, pady=(0, 15))
        
        # Description field
        desc_label = ctk.CTkLabel(
            fields_frame,
            text="Description (Optional):",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        desc_label.pack(anchor="w", padx=15, pady=(0, 5))
        
        self.description_textbox = ctk.CTkTextbox(
            fields_frame,
            height=80,
            font=ctk.CTkFont(size=12)
        )
        self.description_textbox.pack(fill="x", padx=15, pady=(0, 15))
        
        # Fill form if editing
        if sub_category_data:
            self.name_entry.insert(0, sub_category_data.get('sub_category_name', ''))
            if sub_category_data.get('description'):
                self.description_textbox.insert("1.0", sub_category_data['description'])
            
            # Set category
            if sub_category_data.get('category_id'):
                for cat in self.categories:
                    if cat['category_id'] == sub_category_data['category_id']:
                        self.category_id_var.set(cat['category_id'])
                        self.category_display.configure(state="normal")
                        self.category_display.delete(0, "end")
                        self.category_display.insert(0, cat['category_name'])
                        self.category_display.configure(state="readonly")
                        break
        
        # Buttons
        button_frame = ctk.CTkFrame(main_frame)
        button_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        save_btn = ctk.CTkButton(
            button_frame,
            text="Save Sub Category",
            command=self.save_sub_category,
            width=160,
            height=40,
            fg_color=("#8BC34A", "#558B2F"),
            hover_color=("#7CB342", "#33691E"),
            font=ctk.CTkFont(size=13, weight="bold")
        )
        save_btn.pack(side="left", padx=15, pady=15)
        
        cancel_btn = ctk.CTkButton(
            button_frame,
            text="Cancel",
            command=self.cancel,
            width=100,
            height=40,
            fg_color=("#EF4444", "#B91C1C"),
            hover_color=("#DC2626", "#991B1B"),
            font=ctk.CTkFont(size=13, weight="bold")
        )
        cancel_btn.pack(side="right", padx=15, pady=15)
        
        self.dialog.after(100, lambda: self.name_entry.focus())
    
    def show_category_selector(self):
        """Show category selector"""
        selector = SelectionDialog(self.dialog, "Select Category", self.categories, 'category_name', 'category_id')
        if selector.result:
            self.category_id_var.set(selector.result['id'])
            self.category_display.configure(state="normal")
            self.category_display.delete(0, "end")
            self.category_display.insert(0, selector.result['name'])
            self.category_display.configure(state="readonly")
    
    def save_sub_category(self):
        """Save sub category"""
        try:
            name = self.name_entry.get().strip()
            category_id = self.category_id_var.get()
            
            if not name:
                messagebox.showerror("Error", "Sub Category name is required.")
                return
            if not category_id:
                messagebox.showerror("Error", "Please select a parent category.")
                return
            
            description = self.description_textbox.get("1.0", "end-1c").strip()
            
            self.result = {
                'name': name,
                'category_id': category_id,
                'description': description or ""
            }
            
            self.dialog.destroy()
            
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")
    
    def cancel(self):
        """Cancel and close dialog"""
        self.result = None
        self.dialog.destroy()


class BrandDialog:
    def __init__(self, parent, title, brand_data=None):
        self.parent = parent
        self.result = None
        
        # Create dialog
        self.dialog = ctk.CTkToplevel(parent)
        self.dialog.title(title)
        self.dialog.geometry("500x400")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # Center dialog on parent window
        self.dialog.update_idletasks()
        parent.update_idletasks()
        
        parent_x = parent.winfo_rootx()
        parent_y = parent.winfo_rooty()
        parent_width = parent.winfo_width()
        parent_height = parent.winfo_height()
        
        dialog_width = 500
        dialog_height = 400
        
        x = parent_x + (parent_width // 2) - (dialog_width // 2)
        y = parent_y + (parent_height // 2) - (dialog_height // 2)
        
        self.dialog.geometry(f"{dialog_width}x{dialog_height}+{x}+{y}")
        
        self.dialog.attributes("-topmost", True)
        self.dialog.focus_force()
        
        # Create form
        self.create_form(brand_data)
        
        # Wait for dialog to close
        self.dialog.wait_window()
    
    def create_form(self, brand_data):
        """Create brand form with modern styling"""
        main_frame = ctk.CTkScrollableFrame(self.dialog)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Header
        header_frame = ctk.CTkFrame(main_frame, corner_radius=10, fg_color=("#FF9800", "#E65100"))
        header_frame.pack(fill="x", padx=10, pady=(10, 15))
        
        ctk.CTkLabel(
            header_frame,
            text="Brand Information",
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color="white"
        ).pack(pady=15)
        
        # Form fields container
        fields_frame = ctk.CTkFrame(main_frame)
        fields_frame.pack(fill="x", padx=10, pady=(0, 15))
        
        # Name field
        name_label = ctk.CTkLabel(
            fields_frame,
            text="Brand Name:",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        name_label.pack(anchor="w", padx=15, pady=(15, 5))
        
        self.name_entry = ctk.CTkEntry(
            fields_frame,
            placeholder_text="Enter brand name...",
            font=ctk.CTkFont(size=14),
            height=35
        )
        self.name_entry.pack(fill="x", padx=15, pady=(0, 15))
        
        # Description field
        desc_label = ctk.CTkLabel(
            fields_frame,
            text="Description (Optional):",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        desc_label.pack(anchor="w", padx=15, pady=(0, 5))
        
        self.description_textbox = ctk.CTkTextbox(
            fields_frame,
            height=80,
            font=ctk.CTkFont(size=12)
        )
        self.description_textbox.pack(fill="x", padx=15, pady=(0, 15))
        
        # Fill form if editing
        if brand_data:
            self.name_entry.insert(0, brand_data.get('brand_name', ''))
            if brand_data.get('description'):
                self.description_textbox.insert("1.0", brand_data['description'])
        
        # Buttons
        button_frame = ctk.CTkFrame(main_frame)
        button_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        save_btn = ctk.CTkButton(
            button_frame,
            text="Save Brand",
            command=self.save_brand,
            width=140,
            height=40,
            fg_color=("#FF9800", "#E65100"),
            hover_color=("#F57C00", "#BF360C"),
            font=ctk.CTkFont(size=13, weight="bold")
        )
        save_btn.pack(side="left", padx=15, pady=15)
        
        cancel_btn = ctk.CTkButton(
            button_frame,
            text="Cancel",
            command=self.cancel,
            width=100,
            height=40,
            fg_color=("#EF4444", "#B91C1C"),
            hover_color=("#DC2626", "#991B1B"),
            font=ctk.CTkFont(size=13, weight="bold")
        )
        cancel_btn.pack(side="right", padx=15, pady=15)
        
        self.dialog.after(100, lambda: self.name_entry.focus())
    
    def save_brand(self):
        """Save brand"""
        try:
            name = self.name_entry.get().strip()
            if not name:
                messagebox.showerror("Error", "Brand name is required.")
                return
            
            description = self.description_textbox.get("1.0", "end-1c").strip()
            
            self.result = {
                'name': name,
                'description': description or ""
            }
            
            self.dialog.destroy()
            
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")
    
    def cancel(self):
        """Cancel and close dialog"""
        self.result = None
        self.dialog.destroy()


class SubBrandDialog:
    def __init__(self, parent, title, sub_brand_data=None):
        self.parent = parent
        self.result = None
        
        # Create dialog
        self.dialog = ctk.CTkToplevel(parent)
        self.dialog.title(title)
        self.dialog.geometry("500x500")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # Center dialog
        self.dialog.update_idletasks()
        parent.update_idletasks()
        
        parent_x = parent.winfo_rootx()
        parent_y = parent.winfo_rooty()
        parent_width = parent.winfo_width()
        parent_height = parent.winfo_height()
        
        dialog_width = 500
        dialog_height = 500
        
        x = parent_x + (parent_width // 2) - (dialog_width // 2)
        y = parent_y + (parent_height // 2) - (dialog_height // 2)
        
        self.dialog.geometry(f"{dialog_width}x{dialog_height}+{x}+{y}")
        
        self.dialog.attributes("-topmost", True)
        self.dialog.focus_force()
        
        # Load brands
        self.brands = db.execute_query("SELECT brand_id, brand_name FROM brands ORDER BY brand_name")
        
        # Create form
        self.create_form(sub_brand_data)
        
        # Wait for dialog to close
        self.dialog.wait_window()
    
    def create_form(self, sub_brand_data):
        """Create sub brand form"""
        main_frame = ctk.CTkScrollableFrame(self.dialog)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Header
        header_frame = ctk.CTkFrame(main_frame, corner_radius=10, fg_color=("#FFB74D", "#EF6C00"))
        header_frame.pack(fill="x", padx=10, pady=(10, 15))
        
        ctk.CTkLabel(
            header_frame,
            text="Sub Brand Information",
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color="white"
        ).pack(pady=15)
        
        # Form fields
        fields_frame = ctk.CTkFrame(main_frame)
        fields_frame.pack(fill="x", padx=10, pady=(0, 15))
        
        # Brand selection
        brand_label = ctk.CTkLabel(
            fields_frame,
            text="Parent Brand:*",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        brand_label.pack(anchor="w", padx=15, pady=(15, 5))
        
        self.brand_var = tk.StringVar()
        self.brand_id_var = tk.IntVar()
        brand_frame = ctk.CTkFrame(fields_frame, fg_color="transparent")
        brand_frame.pack(fill="x", padx=15, pady=(0, 15))
        
        self.brand_display = ctk.CTkEntry(brand_frame, width=320, state="readonly", font=ctk.CTkFont(size=12))
        self.brand_display.pack(side="left", padx=(0, 10))
        
        ctk.CTkButton(brand_frame, text="Search", width=70, command=self.show_brand_selector).pack(side="left")
        
        # Name field
        name_label = ctk.CTkLabel(
            fields_frame,
            text="Sub Brand Name:*",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        name_label.pack(anchor="w", padx=15, pady=(0, 5))
        
        self.name_entry = ctk.CTkEntry(
            fields_frame,
            placeholder_text="Enter sub brand name...",
            font=ctk.CTkFont(size=14),
            height=35
        )
        self.name_entry.pack(fill="x", padx=15, pady=(0, 15))
        
        # Description field
        desc_label = ctk.CTkLabel(
            fields_frame,
            text="Description (Optional):",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        desc_label.pack(anchor="w", padx=15, pady=(0, 5))
        
        self.description_textbox = ctk.CTkTextbox(
            fields_frame,
            height=80,
            font=ctk.CTkFont(size=12)
        )
        self.description_textbox.pack(fill="x", padx=15, pady=(0, 15))
        
        # Fill form if editing
        if sub_brand_data:
            self.name_entry.insert(0, sub_brand_data.get('sub_brand_name', ''))
            if sub_brand_data.get('description'):
                self.description_textbox.insert("1.0", sub_brand_data['description'])
            
            # Set brand
            if sub_brand_data.get('brand_id'):
                for brand in self.brands:
                    if brand['brand_id'] == sub_brand_data['brand_id']:
                        self.brand_id_var.set(brand['brand_id'])
                        self.brand_display.configure(state="normal")
                        self.brand_display.delete(0, "end")
                        self.brand_display.insert(0, brand['brand_name'])
                        self.brand_display.configure(state="readonly")
                        break
        
        # Buttons
        button_frame = ctk.CTkFrame(main_frame)
        button_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        save_btn = ctk.CTkButton(
            button_frame,
            text="Save Sub Brand",
            command=self.save_sub_brand,
            width=160,
            height=40,
            fg_color=("#FFB74D", "#EF6C00"),
            hover_color=("#FFA726", "#E65100"),
            font=ctk.CTkFont(size=13, weight="bold")
        )
        save_btn.pack(side="left", padx=15, pady=15)
        
        cancel_btn = ctk.CTkButton(
            button_frame,
            text="Cancel",
            command=self.cancel,
            width=100,
            height=40,
            fg_color=("#EF4444", "#B91C1C"),
            hover_color=("#DC2626", "#991B1B"),
            font=ctk.CTkFont(size=13, weight="bold")
        )
        cancel_btn.pack(side="right", padx=15, pady=15)
        
        self.dialog.after(100, lambda: self.name_entry.focus())
    
    def show_brand_selector(self):
        """Show brand selector"""
        selector = SelectionDialog(self.dialog, "Select Brand", self.brands, 'brand_name', 'brand_id')
        if selector.result:
            self.brand_id_var.set(selector.result['id'])
            self.brand_display.configure(state="normal")
            self.brand_display.delete(0, "end")
            self.brand_display.insert(0, selector.result['name'])
            self.brand_display.configure(state="readonly")
    
    def save_sub_brand(self):
        """Save sub brand"""
        try:
            name = self.name_entry.get().strip()
            brand_id = self.brand_id_var.get()
            
            if not name:
                messagebox.showerror("Error", "Sub Brand name is required.")
                return
            if not brand_id:
                messagebox.showerror("Error", "Please select a parent brand.")
                return
            
            description = self.description_textbox.get("1.0", "end-1c").strip()
            
            self.result = {
                'name': name,
                'brand_id': brand_id,
                'description': description or ""
            }
            
            self.dialog.destroy()
            
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")
    
    def cancel(self):
        """Cancel and close dialog"""
        self.result = None
        self.dialog.destroy()


class SearchableDropdown(ctk.CTkFrame):
    """Custom searchable dropdown component combining dropdown and search"""
    def __init__(self, parent, items, name_field, id_field, placeholder="Select...", 
                 width=400, height=35, command=None, add_command=None, **kwargs):
        super().__init__(parent, fg_color="transparent", **kwargs)
        
        self.items = items
        self.name_field = name_field
        self.id_field = id_field
        self.command = command
        self.add_command = add_command
        self.selected_id = None
        self.selected_name = None
        self.dropdown_open = False
        
        # Main container frame
        self.container = ctk.CTkFrame(self, fg_color=("white", "#2D2D2D"), 
                                       border_color=("#D1D5DB", "#404040"),
                                       border_width=1, corner_radius=6)
        self.container.pack(fill="both", expand=True)
        
        # Display label (shows selected value)
        self.display_var = tk.StringVar(value=placeholder)
        self.display_label = ctk.CTkLabel(self.container, textvariable=self.display_var,
                                          font=ctk.CTkFont(size=12),
                                          text_color=("gray40", "gray70"),
                                          width=width-60, anchor="w")
        self.display_label.pack(side="left", padx=(12, 5), pady=5)
        
        # Dropdown button with chevron
        self.dropdown_btn = ctk.CTkButton(self.container, text="▼", width=30, height=height-4,
                                          fg_color="transparent", text_color=("gray40", "gray70"),
                                          hover_color=("gray85", "gray35"),
                                          command=self.toggle_dropdown)
        self.dropdown_btn.pack(side="right", padx=(0, 3), pady=3)
        
        # Make the entire container clickable
        self.container.bind("<Button-1>", lambda e: self.toggle_dropdown())
        self.display_label.bind("<Button-1>", lambda e: self.toggle_dropdown())
        
        # Dropdown window (created when needed)
        self.dropdown_window = None
        self.search_var = None
        self.listbox = None
        self.filtered_items = items.copy()
    
    def set_items(self, items):
        """Update the items in the dropdown"""
        self.items = items
        self.filtered_items = items.copy()
    
    def get(self):
        """Get the selected value"""
        return self.selected_id
    
    def get_name(self):
        """Get the selected name"""
        return self.selected_name
    
    def set(self, name, id_value):
        """Set the selected value"""
        self.selected_name = name
        self.selected_id = id_value
        self.display_var.set(name)
        self.display_label.configure(text_color=("gray10", "gray90"))
    
    def clear(self):
        """Clear the selection"""
        self.selected_name = None
        self.selected_id = None
        self.display_var.set("Select...")
        self.display_label.configure(text_color=("gray40", "gray70"))
    
    def toggle_dropdown(self):
        """Toggle the dropdown window"""
        if self.dropdown_open:
            self.close_dropdown()
        else:
            self.open_dropdown()
    
    def open_dropdown(self):
        """Open the dropdown window with search and list"""
        if self.dropdown_open:
            return
        
        self.dropdown_open = True
        self.dropdown_btn.configure(text="▲")
        
        # Create dropdown window
        self.dropdown_window = ctk.CTkToplevel(self.winfo_toplevel())
        self.dropdown_window.withdraw()
        self.dropdown_window.overrideredirect(True)  # Remove window decorations
        self.dropdown_window.attributes("-topmost", True)
        
        # Dropdown frame
        dropdown_frame = ctk.CTkFrame(self.dropdown_window, fg_color=("white", "#2D2D2D"),
                                       border_color=("#D1D5DB", "#404040"),
                                       border_width=1, corner_radius=6)
        dropdown_frame.pack(fill="both", expand=True, padx=1, pady=1)
        
        # Search frame
        search_frame = ctk.CTkFrame(dropdown_frame, fg_color="transparent")
        search_frame.pack(fill="x", padx=10, pady=(10, 5))
        
        # Search icon
        search_icon = ctk.CTkLabel(search_frame, text="🔍", font=ctk.CTkFont(size=12))
        search_icon.pack(side="left", padx=(0, 5))
        
        # Search entry
        self.search_var = tk.StringVar()
        self.search_var.trace('w', self._filter_items)
        search_entry = ctk.CTkEntry(search_frame, textvariable=self.search_var, 
                                    placeholder_text="Search...",
                                    font=ctk.CTkFont(size=12),
                                    height=32, border_width=0)
        search_entry.pack(side="left", fill="x", expand=True)
        search_entry.focus()
        
        # Separator
        separator = ctk.CTkFrame(dropdown_frame, height=1, fg_color=("#E5E7EB", "#404040"))
        separator.pack(fill="x", padx=10, pady=5)
        
        if hasattr(self, 'add_command') and self.add_command:
            add_btn = ctk.CTkButton(dropdown_frame, text="+ Add New", width=100, height=28,
                                    fg_color=("#3B82F6", "#2563EB"), hover_color=("#2563EB", "#1D4ED8"),
                                    font=ctk.CTkFont(size=11), command=self._on_add_new)
            add_btn.pack(anchor="w", padx=10, pady=(0, 5))
            
            separator2 = ctk.CTkFrame(dropdown_frame, height=1, fg_color=("#E5E7EB", "#404040"))
            separator2.pack(fill="x", padx=10, pady=5)
        
        # List frame
        list_frame = ctk.CTkFrame(dropdown_frame, fg_color="transparent")
        list_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        
        # Scrollbar
        scrollbar = ctk.CTkScrollbar(list_frame)
        scrollbar.pack(side="right", fill="y")
        
        is_dark = ctk.get_appearance_mode() == "Dark"
        self.listbox = tk.Listbox(list_frame, font=("Segoe UI", 12), 
                                   bg="#2D2D2D" if is_dark else "#FFFFFF",
                                   fg="#E0E0E0" if is_dark else "#1F2937",
                                   selectbackground=("#3B82F6", "#2563EB")[is_dark],
                                   selectforeground="white",
                                   borderwidth=0, highlightthickness=0,
                                   activestyle="none",
                                   yscrollcommand=scrollbar.set)
        self.listbox.pack(side="left", fill="both", expand=True)
        scrollbar.configure(command=self.listbox.yview)
        
        # Bind events
        self.listbox.bind("<Double-1>", lambda e: self._select_item())
        self.listbox.bind("<Return>", lambda e: self._select_item())
        self.listbox.bind("<Button-1>", lambda e: self._on_list_click(e))
        
        # Close on escape or click outside
        self.dropdown_window.bind("<Escape>", lambda e: self.close_dropdown())
        self.dropdown_window.bind("<FocusOut>", lambda e: self._on_focus_out())
        
        # Populate list
        self._populate_list()
        
        # Bind click outside to close
        self.winfo_toplevel().bind("<Button-1>", self._on_click_outside, add="+")
        
        # Position and show
        x = self.container.winfo_rootx()
        y = self.container.winfo_rooty() + self.container.winfo_height() + 2
        width = self.container.winfo_width()
        self.dropdown_window.geometry(f"{width}x350+{x}+{y}")
        self.dropdown_window.deiconify()
        self.dropdown_window.lift()
        self.dropdown_window.focus_force()
    
    def close_dropdown(self):
        """Close the dropdown window"""
        if self.dropdown_window:
            self.dropdown_window.destroy()
            self.dropdown_window = None
        self.dropdown_open = False
        self.dropdown_btn.configure(text="▼")
        self.winfo_toplevel().unbind("<Button-1>")
    
    def _on_click_outside(self, event):
        """Handle click outside dropdown"""
        if self.dropdown_window:
            # Check if click is outside dropdown
            x = self.dropdown_window.winfo_rootx()
            y = self.dropdown_window.winfo_rooty()
            width = self.dropdown_window.winfo_width()
            height = self.dropdown_window.winfo_height()
            
            if not (x <= event.x_root <= x + width and y <= event.y_root <= y + height):
                # Also check if click is on the container
                cx = self.container.winfo_rootx()
                cy = self.container.winfo_rooty()
                cw = self.container.winfo_width()
                ch = self.container.winfo_height()
                
                if not (cx <= event.x_root <= cx + cw and cy <= event.y_root <= cy + ch):
                    self.close_dropdown()
    
    def _on_focus_out(self):
        """Handle focus out event"""
        self.winfo_toplevel().after(100, self._check_focus)
    
    def _check_focus(self):
        """Check if focus is still in dropdown"""
        if self.dropdown_window and not self.dropdown_window.focus_get():
            self.close_dropdown()
    
    def _populate_list(self):
        """Populate the listbox with items"""
        self.listbox.delete(0, tk.END)
        for item in self.filtered_items:
            self.listbox.insert(tk.END, item[self.name_field])
        
        # Highlight selected item
        if self.selected_name:
            for i, item in enumerate(self.filtered_items):
                if item[self.name_field] == self.selected_name:
                    self.listbox.selection_set(i)
                    self.listbox.see(i)
                    break
    
    def _filter_items(self, *args):
        """Filter items based on search"""
        search_term = self.search_var.get().lower().strip()
        
        if not search_term:
            self.filtered_items = self.items.copy()
        else:
            self.filtered_items = [item for item in self.items 
                                   if search_term in item[self.name_field].lower()]
        
        self._populate_list()
    
    def _on_list_click(self, event):
        """Handle list click"""
        # Get the item under cursor
        index = self.listbox.nearest(event.y)
        if 0 <= index < self.listbox.size():
            self.listbox.selection_clear(0, tk.END)
            self.listbox.selection_set(index)
    
    def _select_item(self):
        """Select the highlighted item"""
        selection = self.listbox.curselection()
        if not selection:
            return
        
        selected_name = self.listbox.get(selection[0])
        
        # Find the item with matching name
        for item in self.items:
            if item[self.name_field] == selected_name:
                self.set(selected_name, item[self.id_field])
                break
        
        self.close_dropdown()
        
        # Call callback if provided
        if self.command:
            self.command()
    
    def _on_add_new(self):
        """Handle Add New button click"""
        self.close_dropdown()
        if self.add_command:
            self.add_command()


class SelectionDialog:
    """Searchable selection dialog for categories and brands"""
    def __init__(self, parent, title, items, name_field, id_field):
        self.parent = parent
        self.items = items
        self.name_field = name_field
        self.id_field = id_field
        self.result = None
        
        # Create dialog
        self.dialog = ctk.CTkToplevel(parent)
        self.dialog.title(title)
        self.dialog.geometry("400x500")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # Center dialog
        self.dialog.update_idletasks()
        x = parent.winfo_x() + (parent.winfo_width() // 2) - (400 // 2)
        y = parent.winfo_y() + (parent.winfo_height() // 2) - (500 // 2)
        self.dialog.geometry(f"400x500+{x}+{y}")
        
        self.create_ui()
        self.dialog.wait_window()
    
    def create_ui(self):
        """Create the selection UI"""
        is_dark = ctk.get_appearance_mode() == "Dark"
        
        # Header
        header = ctk.CTkFrame(self.dialog, corner_radius=0, fg_color=("#3B82F6", "#2563EB"))
        header.pack(fill="x", pady=0)
        ctk.CTkLabel(header, text=self.dialog.title(), font=ctk.CTkFont(size=16, weight="bold"), text_color="white").pack(pady=15)
        
        # Search frame
        search_frame = ctk.CTkFrame(self.dialog, fg_color="transparent")
        search_frame.pack(fill="x", padx=20, pady=15)
        
        ctk.CTkLabel(search_frame, text="Search:", font=ctk.CTkFont(size=12)).pack(side="left", padx=(0, 10))
        self.search_var = ctk.StringVar()
        self.search_var.trace('w', self.filter_items)
        search_entry = ctk.CTkEntry(search_frame, textvariable=self.search_var, width=280, placeholder_text="Type to search...")
        search_entry.pack(side="left")
        
        # List frame
        list_frame = ctk.CTkFrame(self.dialog)
        list_frame.pack(fill="both", expand=True, padx=20, pady=(0, 15))
        
        self.listbox = tk.Listbox(list_frame, font=("Segoe UI", 12),
                                  bg="#2D2D2D" if is_dark else "#FFFFFF",
                                  fg="#E0E0E0" if is_dark else "#1F2937",
                                  selectbackground=("#3B82F6", "#2563EB")[is_dark],
                                  selectforeground="white",
                                  borderwidth=0, highlightthickness=0)
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=self.listbox.yview)
        self.listbox.configure(yscrollcommand=scrollbar.set)
        
        self.listbox.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        self.listbox.bind("<Double-1>", lambda e: self.select_item())
        self.populate_list()
        
        # Buttons
        button_frame = ctk.CTkFrame(self.dialog, fg_color="transparent")
        button_frame.pack(fill="x", padx=20, pady=(0, 20))
        
        ctk.CTkButton(button_frame, text="Select", command=self.select_item,
                     fg_color=("#3B82F6", "#2563EB"), hover_color=("#2563EB", "#1D4ED8"),
                     width=100).pack(side="left", padx=(0, 10))
        ctk.CTkButton(button_frame, text="Cancel", command=self.cancel,
                     fg_color=("#6B7280", "#4B5563"), hover_color=("#4B5563", "#374151"),
                     width=100).pack(side="left")
    
    def populate_list(self):
        """Populate the listbox with items"""
        self.listbox.delete(0, tk.END)
        for item in self.items:
            self.listbox.insert(tk.END, item[self.name_field])
    
    def filter_items(self, *args):
        """Filter items based on search"""
        search_term = self.search_var.get().lower()
        self.listbox.delete(0, tk.END)
        
        for item in self.items:
            name = item[self.name_field].lower()
            if search_term in name:
                self.listbox.insert(tk.END, item[self.name_field])
    
    def select_item(self):
        """Select the highlighted item"""
        selection = self.listbox.curselection()
        if not selection:
            messagebox.showwarning("Warning", "Please select an item")
            return
        
        selected_name = self.listbox.get(selection[0])
        
        # Find the item with matching name
        for item in self.items:
            if item[self.name_field] == selected_name:
                self.result = {'id': item[self.id_field], 'name': selected_name}
                break
        
        self.dialog.destroy()
    
    def cancel(self):
        """Cancel selection"""
        self.result = None
        self.dialog.destroy()
