import os
import sys
from datetime import datetime
from tkinter import messagebox, filedialog

# Add parent directory to path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from models.database import db

class ExportManager:
    """Manager class for exporting data to Excel and PDF formats"""
    
    @staticmethod
    def export_to_excel(data, headers, filename, title=None):
        """
        Export data to Excel file
        
        Args:
            data: List of dictionaries or sqlite3.Row objects
            headers: List of column headers
            filename: Suggested filename
            title: Optional title for the sheet
        """
        try:
            # Try to import openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                from openpyxl.utils import get_column_letter
            except ImportError:
                messagebox.showerror(
                    "Missing Dependency",
                    "openpyxl is required for Excel export.\n\nInstall it with:\npip install openpyxl"
                )
                return False
            
            # Ask user for save location
            file_path = filedialog.asksaveasfilename(
                title="Export to Excel",
                initialfile=filename,
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if not file_path:
                return False
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            if title:
                ws.title = title[:31]  # Excel sheet name max length is 31
            
            # Add title row if provided
            start_row = 1
            if title:
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
                title_cell = ws.cell(row=1, column=1, value=title)
                title_cell.font = Font(size=16, bold=True, color="FFFFFF")
                title_cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[1].height = 30
                start_row = 2
            
            # Add headers
            header_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
            header_font = Font(bold=True, size=11)
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Add data - access by index since sqlite3.Row doesn't match header names
            for row_idx, row_data in enumerate(data, start_row + 1):
                # Get the actual column names from the row
                if hasattr(row_data, 'keys'):
                    columns = list(row_data.keys())
                else:
                    columns = None
                
                for col_idx, header in enumerate(headers, 1):
                    # Access by index if we have columns, otherwise try direct access
                    if columns and col_idx <= len(columns):
                        col_name = columns[col_idx - 1]
                        value = row_data[col_name]
                    elif hasattr(row_data, '__getitem__'):
                        try:
                            value = row_data[col_idx - 1]
                        except (IndexError, KeyError):
                            value = ''
                    else:
                        value = ''
                    
                    cell = ws.cell(row=row_idx, column=col_idx, value=value if value is not None else '')
                    cell.alignment = Alignment(vertical="center")
            
            # Auto-adjust column widths
            for col_idx, header in enumerate(headers, 1):
                max_length = len(str(header))
                for row_data in data:
                    if hasattr(row_data, 'keys'):
                        columns = list(row_data.keys())
                        if col_idx <= len(columns):
                            col_name = columns[col_idx - 1]
                            value = row_data[col_name]
                        else:
                            value = ''
                    elif hasattr(row_data, '__getitem__'):
                        try:
                            value = row_data[col_idx - 1]
                        except (IndexError, KeyError):
                            value = ''
                    else:
                        value = ''
                    max_length = max(max_length, len(str(value)) if value is not None else 0)
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
            
            # Save workbook
            wb.save(file_path)
            
            messagebox.showinfo(
                "Export Successful",
                f"✅ Data exported successfully!\n\nSaved to:\n{file_path}"
            )
            return True
            
        except Exception as e:
            messagebox.showerror("Export Failed", f"Failed to export to Excel:\n{e}")
            return False
    
    @staticmethod
    def export_to_pdf(data, headers, filename, title=None, subtitle=None):
        """
        Export data to PDF file
        
        Args:
            data: List of dictionaries or sqlite3.Row objects
            headers: List of column headers
            filename: Suggested filename
            title: Optional title for the document
            subtitle: Optional subtitle
        """
        try:
            # Try to import reportlab
            try:
                from reportlab.lib import colors
                from reportlab.lib.pagesizes import letter, A4
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import inch
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.enums import TA_CENTER, TA_LEFT
            except ImportError:
                messagebox.showerror(
                    "Missing Dependency",
                    "reportlab is required for PDF export.\n\nInstall it with:\npip install reportlab"
                )
                return False
            
            # Ask user for save location
            file_path = filedialog.asksaveasfilename(
                title="Export to PDF",
                initialfile=filename,
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
            )
            
            if not file_path:
                return False
            
            # Create PDF document
            doc = SimpleDocTemplate(
                file_path,
                pagesize=A4,
                rightMargin=0.5*inch,
                leftMargin=0.5*inch,
                topMargin=0.5*inch,
                bottomMargin=0.5*inch
            )
            
            # Container for elements
            elements = []
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#3B82F6'),
                spaceAfter=12,
                alignment=TA_CENTER
            )
            subtitle_style = ParagraphStyle(
                'CustomSubtitle',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.gray,
                spaceAfter=20,
                alignment=TA_CENTER
            )
            
            # Add title
            if title:
                elements.append(Paragraph(title, title_style))
            
            # Add subtitle with date
            if subtitle:
                elements.append(Paragraph(subtitle, subtitle_style))
            
            date_text = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            elements.append(Paragraph(date_text, subtitle_style))
            elements.append(Spacer(1, 0.2*inch))
            
            # Prepare table data
            table_data = [headers]
            
            for row_data in data:
                row = []
                # Get the actual column names from the row
                if hasattr(row_data, 'keys'):
                    columns = list(row_data.keys())
                else:
                    columns = None
                
                for col_idx, header in enumerate(headers, 1):
                    # Access by index if we have columns
                    if columns and col_idx <= len(columns):
                        col_name = columns[col_idx - 1]
                        value = row_data[col_name]
                    elif hasattr(row_data, '__getitem__'):
                        try:
                            value = row_data[col_idx - 1]
                        except (IndexError, KeyError):
                            value = ''
                    else:
                        value = ''
                    row.append(str(value) if value is not None else '')
                table_data.append(row)
            
            # Create table
            table = Table(table_data, repeatRows=1)
            
            # Style the table
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ])
            table.setStyle(table_style)
            
            elements.append(table)
            
            # Build PDF
            doc.build(elements)
            
            messagebox.showinfo(
                "Export Successful",
                f"✅ Data exported successfully!\n\nSaved to:\n{file_path}"
            )
            return True
            
        except Exception as e:
            messagebox.showerror("Export Failed", f"Failed to export to PDF:\n{e}")
            return False
    
    @staticmethod
    def export_inventory_to_excel(filter_category=None, filter_brand=None, filter_stock=None):
        """Export inventory data to Excel"""
        query = """
            SELECT p.sku, p.name, c.category_name, b.brand_name, p.stock,
                   p.cost_price, p.price_normal, p.price_workshop, p.reorder_level
            FROM products p
            LEFT JOIN categories c ON p.category_id = c.category_id
            LEFT JOIN brands b ON p.brand_id = b.brand_id
            WHERE p.is_active = 1
        """
        params = []
        
        if filter_category and filter_category != "All":
            query += " AND c.category_name = ?"
            params.append(filter_category)
        
        if filter_brand and filter_brand != "All":
            query += " AND b.brand_name = ?"
            params.append(filter_brand)
        
        if filter_stock and filter_stock != "All":
            if filter_stock == "Low Stock":
                query += " AND p.stock <= p.reorder_level AND p.stock > 0"
            elif filter_stock == "Out of Stock":
                query += " AND p.stock = 0"
            elif filter_stock == "In Stock":
                query += " AND p.stock > p.reorder_level"
        
        query += " ORDER BY p.name"
        
        products = db.execute_query(query, params if params else None)
        
        headers = ["SKU", "Product Name", "Category", "Brand", "Stock", 
                   "Cost Price", "Normal Price", "Workshop Price", "Reorder Level"]
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Inventory_Report_{timestamp}.xlsx"
        title = "Inventory Report"
        
        return ExportManager.export_to_excel(products, headers, filename, title)
    
    @staticmethod
    def export_inventory_to_pdf(filter_category=None, filter_brand=None, filter_stock=None):
        """Export inventory data to PDF"""
        query = """
            SELECT p.sku, p.name, c.category_name, b.brand_name, p.stock,
                   p.cost_price, p.price_normal, p.reorder_level
            FROM products p
            LEFT JOIN categories c ON p.category_id = c.category_id
            LEFT JOIN brands b ON p.brand_id = b.brand_id
            WHERE p.is_active = 1
        """
        params = []
        
        if filter_category and filter_category != "All":
            query += " AND c.category_name = ?"
            params.append(filter_category)
        
        if filter_brand and filter_brand != "All":
            query += " AND b.brand_name = ?"
            params.append(filter_brand)
        
        if filter_stock and filter_stock != "All":
            if filter_stock == "Low Stock":
                query += " AND p.stock <= p.reorder_level AND p.stock > 0"
            elif filter_stock == "Out of Stock":
                query += " AND p.stock = 0"
            elif filter_stock == "In Stock":
                query += " AND p.stock > p.reorder_level"
        
        query += " ORDER BY p.name"
        
        products = db.execute_query(query, params if params else None)
        
        headers = ["SKU", "Product Name", "Category", "Brand", "Stock", 
                   "Cost Price", "Normal Price", "Reorder Level"]
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Inventory_Report_{timestamp}.pdf"
        title = "Inventory Report"
        
        # Build subtitle with filter info
        filters = []
        if filter_category and filter_category != "All":
            filters.append(f"Category: {filter_category}")
        if filter_brand and filter_brand != "All":
            filters.append(f"Brand: {filter_brand}")
        if filter_stock and filter_stock != "All":
            filters.append(f"Stock: {filter_stock}")
        
        subtitle = " | ".join(filters) if filters else "All Products"
        
        return ExportManager.export_to_pdf(products, headers, filename, title, subtitle)
    
    @staticmethod
    def export_customers_to_excel():
        """Export customers data to Excel"""
        query = """
            SELECT customer_id, name, type, contact, credit_balance, is_active, created_at
            FROM customers
            ORDER BY name
        """
        customers = db.execute_query(query)
        
        headers = ["ID", "Name", "Type", "Contact", "Credit Balance", "Active", "Created Date"]
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Customers_Report_{timestamp}.xlsx"
        title = "Customers Report"
        
        return ExportManager.export_to_excel(customers, headers, filename, title)
    
    @staticmethod
    def export_customers_to_pdf():
        """Export customers data to PDF"""
        query = """
            SELECT customer_id, name, type, contact, credit_balance, is_active
            FROM customers
            ORDER BY name
        """
        customers = db.execute_query(query)
        
        headers = ["ID", "Name", "Type", "Contact", "Credit Balance", "Active"]
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Customers_Report_{timestamp}.pdf"
        title = "Customers Report"
        
        return ExportManager.export_to_pdf(customers, headers, filename, title)
    
    @staticmethod
    def export_suppliers_to_excel():
        """Export suppliers data to Excel"""
        query = """
            SELECT supplier_id, name, contact_person, email, phone, address, credit_balance, is_active
            FROM suppliers
            ORDER BY name
        """
        suppliers = db.execute_query(query)
        
        headers = ["ID", "Name", "Contact Person", "Email", "Phone", "Address", "Credit Balance", "Active"]
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Suppliers_Report_{timestamp}.xlsx"
        title = "Suppliers Report"
        
        return ExportManager.export_to_excel(suppliers, headers, filename, title)
    
    @staticmethod
    def export_suppliers_to_pdf():
        """Export suppliers data to PDF"""
        query = """
            SELECT supplier_id, name, contact_person, email, phone, credit_balance
            FROM suppliers
            ORDER BY name
        """
        suppliers = db.execute_query(query)
        
        headers = ["ID", "Name", "Contact Person", "Email", "Phone", "Credit Balance"]
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Suppliers_Report_{timestamp}.pdf"
        title = "Suppliers Report"
        
        return ExportManager.export_to_pdf(suppliers, headers, filename, title)
    
    @staticmethod
    def export_sales_to_excel(start_date=None, end_date=None):
        """Export sales data to Excel"""
        query = """
            SELECT s.invoice_number, s.sale_date, c.name as customer_name,
                   s.total_amount, s.paid_amount, s.balance, s.payment_method, s.status
            FROM sales s
            LEFT JOIN customers c ON s.customer_id = c.customer_id
            WHERE 1=1
        """
        params = []
        
        if start_date:
            query += " AND s.sale_date >= ?"
            params.append(start_date)
        if end_date:
            query += " AND s.sale_date <= ?"
            params.append(end_date)
        
        query += " ORDER BY s.sale_date DESC"
        
        sales = db.execute_query(query, params if params else None)
        
        headers = ["Invoice #", "Date", "Customer", "Total Amount", "Paid", "Balance", "Payment Method", "Status"]
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Sales_Report_{timestamp}.xlsx"
        title = "Sales Report"
        
        return ExportManager.export_to_excel(sales, headers, filename, title)
    
    @staticmethod
    def export_sales_to_pdf(start_date=None, end_date=None):
        """Export sales data to PDF"""
        query = """
            SELECT s.invoice_number, s.sale_date, c.name as customer_name,
                   s.total_amount, s.paid_amount, s.balance, s.status
            FROM sales s
            LEFT JOIN customers c ON s.customer_id = c.customer_id
            WHERE 1=1
        """
        params = []
        
        if start_date:
            query += " AND s.sale_date >= ?"
            params.append(start_date)
        if end_date:
            query += " AND s.sale_date <= ?"
            params.append(end_date)
        
        query += " ORDER BY s.sale_date DESC"
        
        sales = db.execute_query(query, params if params else None)
        
        headers = ["Invoice #", "Date", "Customer", "Total", "Paid", "Balance", "Status"]
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Sales_Report_{timestamp}.pdf"
        title = "Sales Report"
        
        subtitle = None
        if start_date and end_date:
            subtitle = f"Period: {start_date} to {end_date}"
        elif start_date:
            subtitle = f"From: {start_date}"
        elif end_date:
            subtitle = f"Until: {end_date}"
        
        return ExportManager.export_to_pdf(sales, headers, filename, title, subtitle)